"""
test_prn_only.py — Prueba la generación PRN con factura simulada.
Usa una factura de ejemplo con referencias reales del Excel.
"""

import sys
from pathlib import Path
import openpyxl
from dotenv import load_dotenv

load_dotenv()

EXCEL_PATH     = Path("CARGUE FACTURAS DE COMPRA F3 ACT.xlsm")
NIT_INCOLMOTOS = "890916911"

# ─── PASO 1: CARGAR EXCEL ─────────────────────────────────────────────────────

print("=" * 60)
print("PASO 1: Cargando Excel...")
print("=" * 60)

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, keep_vba=True)

invenarios: dict[str, dict] = {}
ws_inv = wb["INVENARIOS"]
for row in ws_inv.iter_rows(min_row=3, values_only=True):
    ref = row[1]
    if ref is None or str(ref).strip() == "":
        continue
    invenarios[str(ref).strip()] = {
        "producto": str(row[2]).strip() if row[2] is not None else "",
        "cta_inv":  str(row[4]).strip() if row[4] is not None else "",
    }

datos_tiendas: dict[str, dict] = {}
ws_dat = wb["DATOS"]
for row in ws_dat.iter_rows(min_row=2, values_only=True):
    tienda = row[1]
    if tienda is None or str(tienda).strip() == "":
        continue
    cc  = row[2]
    doc = row[3]
    datos_tiendas[str(tienda).upper().strip()] = {
        "cc":  int(cc)  if cc  is not None else 0,
        "doc": int(doc) if doc is not None else 0,
    }

print(f"  OK: {len(invenarios)} referencias | {len(datos_tiendas)} tiendas")
print(f"  Tiendas: {list(datos_tiendas.keys())}")

# Mostrar las primeras 3 referencias reales del Excel
refs_muestra = list(invenarios.items())[:3]
print("\n  Primeras 3 referencias del Excel:")
for ref, data in refs_muestra:
    print(f"    {ref} | cta={data['cta_inv']} | prod={data['producto'][:30]}")

# ─── FACTURA SIMULADA con referencias reales ──────────────────────────────────

print()
print("=" * 60)
print("PASO 2: Usando factura simulada con referencias reales...")
print("=" * 60)

# Tomamos las 3 primeras referencias reales del Excel
refs_reales = [(ref, data) for ref, data in list(invenarios.items())[:3]]

factura = {
    "numero_factura": "CPFE-750304",
    "fecha": "2025-03-05",
    "ciudad": "NEIVA",
    "subtotal": 5000000.0,
    "iva_total": 950000.0,
    "items": [
        {
            "referencia":   refs_reales[0][0],
            "descripcion":  refs_reales[0][1]["producto"][:50],
            "cantidad":     1,
            "valor_total":  2000000.0,
            "tiene_iva":    True,
        },
        {
            "referencia":   refs_reales[1][0],
            "descripcion":  refs_reales[1][1]["producto"][:50],
            "cantidad":     2,
            "valor_total":  2000000.0,
            "tiene_iva":    True,
        },
        {
            "referencia":   refs_reales[2][0],
            "descripcion":  refs_reales[2][1]["producto"][:50],
            "cantidad":     1,
            "valor_total":  1000000.0,
            "tiene_iva":    False,
        },
    ],
}

print(f"  Factura: {factura['numero_factura']}")
print(f"  Ciudad:  {factura['ciudad']}")
print(f"  Items:   {len(factura['items'])}")

# ─── PASO 3: VALIDAR REFERENCIAS ─────────────────────────────────────────────

print()
print("=" * 60)
print("PASO 3: Validando referencias...")
print("=" * 60)

for item in factura["items"]:
    ref = str(item["referencia"]).strip()
    if ref in invenarios:
        print(f"  OK: {ref} -> cta={invenarios[ref]['cta_inv']}")
    else:
        print(f"  FALTA: {ref}")
        sys.exit(1)

# ─── PASO 4: GENERAR PRN ─────────────────────────────────────────────────────

print()
print("=" * 60)
print("PASO 4: Generando PRN...")
print("=" * 60)

ciudad = factura["ciudad"].upper().strip().split()[0]
cc     = datos_tiendas[ciudad]["cc"]
doc    = datos_tiendas[ciudad]["doc"]
num_doc = factura["numero_factura"].replace("CPFE-", "").replace("CPFE", "").strip()
fecha   = factura["fecha"].replace("-", "")

print(f"  Ciudad={ciudad} CC={cc} DOC={doc} NUM={num_doc} FECHA={fecha}")


def fmt_line(sec, cuenta, producto, descripcion, deb_cred, valor, cantidad):
    line = (
        "P"
        + str(doc).zfill(3)
        + num_doc.zfill(11)
        + str(sec).zfill(5)
        + NIT_INCOLMOTOS.zfill(13)
        + "000"
        + str(cuenta).ljust(10)[:10]
        + str(producto).ljust(13)[:13]
        + fecha
        + str(cc).zfill(4)
        + "000"
        + str(descripcion).ljust(50)[:50]
        + deb_cred
        + str(int(round(valor * 100))).zfill(15)
        + "000000000000000"
        + "0000"
        + "0000"
        + "000"
        + str(cc).zfill(4)
        + "000"
        + str(int(round(cantidad * 100000))).zfill(15)
        + " "
        + "000"
        + "00000000000"
        + "000"
        + "00000000"
        + "0000"
        + "00"
    )
    assert len(line) == 220, f"Linea tiene {len(line)} chars (esperado 220)"
    return line


lines = []
sec = 1

for item in factura["items"]:
    ref  = str(item["referencia"]).strip()
    look = invenarios[ref]
    lines.append(fmt_line(
        sec=sec,
        cuenta=look["cta_inv"],
        producto=look["producto"],
        descripcion=item.get("descripcion", ""),
        deb_cred="D",
        valor=float(item.get("valor_total", 0)),
        cantidad=float(item.get("cantidad", 1)),
    ))
    sec += 1

total_fac = float(factura["subtotal"]) + float(factura["iva_total"])
lines.append(fmt_line(
    sec=sec, cuenta="2205010000", producto="0000000000000",
    descripcion="INCOLMOTOS SAS", deb_cred="C",
    valor=total_fac, cantidad=1,
))
sec += 1

iva = float(factura["iva_total"])
if iva > 0:
    lines.append(fmt_line(
        sec=sec, cuenta="2408020100", producto="0000000000000",
        descripcion="INCOLMOTOS SAS", deb_cred="D",
        valor=iva, cantidad=1,
    ))

# ─── PASO 5: VERIFICAR 220 CHARS ─────────────────────────────────────────────

print()
print("=" * 60)
print("PASO 5: Verificando 220 chars por linea...")
print("=" * 60)

all_ok = True
for i, line in enumerate(lines, 1):
    if len(line) != 220:
        print(f"  ERROR Linea {i}: {len(line)} chars")
        all_ok = False
    else:
        print(f"  OK Linea {i}: 220 chars")

if not all_ok:
    sys.exit(1)

# ─── GUARDAR y MOSTRAR ────────────────────────────────────────────────────────

nombre_prn = f"INT{num_doc}.prn"
content    = "\r\n".join(lines) + "\r\n"
Path(nombre_prn).write_bytes(content.encode("latin-1"))
print(f"\n  Guardado: {nombre_prn} ({len(lines)} lineas)")

print()
print("=" * 60)
print("Primeras 3 lineas del PRN (campos desglosados):")
print("=" * 60)
campos = [
    ("TIPO",        0,   1),
    ("COD.COMP",    1,   4),
    ("NUM.DOC",     4,  15),
    ("SEC",        15,  20),
    ("NIT",        20,  33),
    ("SUCURSAL",   33,  36),
    ("CUENTA",     36,  46),
    ("PRODUCTO",   46,  59),
    ("FECHA",      59,  67),
    ("C.COSTO",    67,  71),
    ("S.COSTO",    71,  74),
    ("DESCRIPCION",74, 124),
    ("DEB.CRED",  124, 125),
    ("VR.MOVIM",  125, 140),
    ("CANTIDAD",  185, 200),
]

for i, line in enumerate(lines[:3], 1):
    print(f"\n  --- Linea {i} ---")
    for nombre, ini, fin in campos:
        print(f"  {nombre:<12}: '{line[ini:fin]}'")

print()
print("=" * 60)
print("PRUEBA END-TO-END COMPLETADA EXITOSAMENTE")
print(f"Archivo generado: {nombre_prn}")
print("=" * 60)
