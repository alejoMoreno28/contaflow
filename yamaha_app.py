"""
yamaha_app.py — ContaFlow Yamaha
Genera archivos .PRN para Siigo desde facturas CPFE de Incolmotos.
"""

import base64
import io
import json
import os
import zipfile
from pathlib import Path

import unicodedata

import anthropic
import openpyxl
import streamlit as st
from dotenv import load_dotenv

load_dotenv()



def _resolver_tienda_ibague(direccion: str):
    import unicodedata
    d = unicodedata.normalize('NFD', str(direccion)).encode('ascii','ignore').decode().upper()
    
    PRINCIPAL_KW = ['CR 5','CRR 5','CARRERA 5','CRA 5','QUINTA','20 39','20-39','B 1 CARMEN','PRINCIPAL']
    SEXTA_KW    = ['CR 6','CRR 6','CARRERA 6','CRA 6','BELALCAZAR','BRR BELALCAZAR','SEXTA','25 40','25-40']
    
    for kw in PRINCIPAL_KW:
        if kw in d:
            return 7, 14
    for kw in SEXTA_KW:
        if kw in d:
            return 1, 1
            
    return 1, 1

def _normalizar_ciudad(texto: str) -> str:
    """Normaliza nombre de ciudad: elimina tildes, mayúsculas, toma primera palabra."""
    return (
        unicodedata.normalize("NFD", texto)
        .encode("ascii", "ignore")
        .decode("utf-8")
        .upper()
        .strip()
        .split()[0]
    )


def _calcular_cta_inv(codigo_producto: str) -> str:
    """
    Calcula cuenta contable desde código producto Siigo.
    Ejemplos confirmados:
    0020080000001 → 1435010280
    0020089001417 → 1435010289
    0020001000001 → 1435010201
    Regla: "14350102" + str(int(codigo[3:7])).zfill(2)
    """
    try:
        sufijo = str(int(codigo_producto[3:7])).zfill(2)
        return "14350102" + sufijo
    except Exception:
        return ""


def _guardar_referencia_en_sheets(referencia: str, producto: str,
                                   descripcion: str, cta_inv: str) -> bool:
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        import json as _json
        from datetime import date

        SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
        SPREADSHEET_ID = "1JzKIDiMmjqVD-iYXTAjxk4wqPvdassNMZJjsNP_UtQI"

        creds_json = os.getenv("GOOGLE_SHEETS_CREDENTIALS")
        if not creds_json:
            try:
                creds_json = st.secrets["GOOGLE_SHEETS_CREDENTIALS"]
            except Exception:
                pass

        if creds_json:
            creds_dict = _json.loads(creds_json)
            creds = Credentials.from_service_account_info(
                creds_dict, scopes=SCOPES
            )
        else:
            creds = Credentials.from_service_account_file(
                "credentials.json", scopes=SCOPES
            )

        gc = gspread.authorize(creds)
        ws = gc.open_by_key(SPREADSHEET_ID).worksheet("INVENARIOS")
        fecha = date.today().strftime("%Y-%m-%d")
        ws.append_row(
            ["", referencia, "'" + str(producto).strip(), descripcion, cta_inv, fecha, "", ""],
            value_input_option="USER_ENTERED"
        )
        return True
    except Exception as e:
        st.error(f"Error guardando en Google Sheets: {e}")
        return False


# ─── CONFIGURACIÓN ────────────────────────────────────────────────────────────

EXCEL_PATH     = Path("CARGUE FACTURAS DE COMPRA F3 ACT.xlsm")
NIT_INCOLMOTOS = "890916911"
GSHEETS_URL    = "https://docs.google.com/spreadsheets/d/1JzKIDiMmjqVD-iYXTAjxk4wqPvdassNMZJjsNP_UtQI/export?format=xlsx"

# ─── PAGE CONFIG ──────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="ContaFlow Yamaha",
    page_icon="🏍️",
    layout="wide",
)

st.markdown("""
<style>
  /* Ocultar elementos nativos de Streamlit */
  #MainMenu {visibility: hidden;}
  footer {visibility: hidden;}
  header {visibility: hidden;}
  
  /* Contenedores Principales */
  .block-container {
      padding-top: 2rem !important;
      padding-bottom: 2rem !important;
      max-width: 1200px !important;
  }
  
  /* Botones Primarios (Rojo Yamaha) */
  .stButton > button[kind="primary"] {
      background-color: #e02828 !important;
      color: white !important;
      border: none !important;
      border-radius: 8px !important;
      font-weight: 600 !important;
      padding: 0.6rem 1.2rem !important;
      transition: all 0.3s ease !important;
      box-shadow: 0 4px 14px 0 rgba(224, 40, 40, 0.39) !important;
  }
  .stButton > button[kind="primary"]:hover {
      transform: translateY(-2px) !important;
      box-shadow: 0 6px 20px rgba(224, 40, 40, 0.5) !important;
      background-color: #ff3333 !important;
  }

  /* Botones Secundarios */
  .stButton > button[kind="secondary"] {
      border-radius: 8px !important;
      border: 1px solid #30363d !important;
      background-color: #161b22 !important;
      color: #c9d1d9 !important;
      transition: all 0.2s ease !important;
  }
  .stButton > button[kind="secondary"]:hover {
      border-color: #8b949e !important;
      background-color: #21262d !important;
  }

  /* Botón de Éxito (Verde PRN) */
  .btn-green > button {
      background-color: #238636 !important;
      color: white !important;
      border: none !important;
      border-radius: 8px !important;
      width: 100% !important;
      font-size: 1.1rem !important;
      font-weight: 600 !important;
      padding: 0.8rem !important;
      transition: all 0.3s ease !important;
      box-shadow: 0 4px 14px 0 rgba(35, 134, 54, 0.39) !important;
  }
  .btn-green > button:hover {
      transform: translateY(-2px) !important;
      background-color: #2ea043 !important;
      box-shadow: 0 6px 20px rgba(35, 134, 54, 0.5) !important;
  }

  /* Cards para Métricas (Subtotal, IVA, Total) */
  div[data-testid="metric-container"] {
      background-color: #161b22;
      border: 1px solid #30363d;
      padding: 1.5rem;
      border-radius: 12px;
      box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
      transition: transform 0.2s ease;
  }
  div[data-testid="metric-container"]:hover {
      transform: translateY(-2px);
      border-color: #8b949e;
  }

  /* Etiquetas de Métricas */
  div[data-testid="stMetricLabel"] > div > div > p {
      color: #8b949e !important;
      font-size: 1rem !important;
      font-weight: 500 !important;
  }

  /* Valores de Métricas */
  div[data-testid="stMetricValue"] > div {
      color: #f0f6fc !important;
      font-weight: 700 !important;
      font-size: 1.8rem !important;
  }

  /* Expanders (Facturas) */
  .streamlit-expanderHeader {
      background-color: #161b22;
      border-radius: 8px !important;
      border: 1px solid #30363d !important;
      padding: 1rem !important;
      font-weight: 600 !important;
      color: #c9d1d9 !important;
  }
  div[data-testid="stExpander"] {
      border: none !important;
      background: transparent !important;
  }

  /* Dataframes / Tables */
  div[data-testid="stDataFrame"] > div {
      border-radius: 8px !important;
      border: 1px solid #30363d !important;
  }
</style>
""", unsafe_allow_html=True)

# ─── VALIDACIÓN API KEY ────────────────────────────────────────────────────────

api_key = os.environ.get("ANTHROPIC_API_KEY", "")
if not api_key:
    try:
        api_key = st.secrets["ANTHROPIC_API_KEY"]
    except Exception:
        pass

if not api_key:
    st.error("⚠️ Error de configuración del sistema. Contacta al administrador.")
    st.stop()

# ─── PASO 1: CARGAR EXCEL AL INICIO ───────────────────────────────────────────

def _parse_wb(wb) -> tuple[dict, dict]:
    invenarios: dict[str, dict] = {}
    ws_inv = wb["INVENARIOS"]
    for row in ws_inv.iter_rows(min_row=3, values_only=True):
        ref = row[1]  # col B
        if ref is None or str(ref).strip() == "":
            continue
        producto = row[2]  # col C
        cta_inv  = row[4]  # col E
        invenarios[str(ref).strip()] = {
            "producto": str(producto).strip() if producto is not None else "",
            "cta_inv":  str(cta_inv).strip()  if cta_inv  is not None else "",
        }

    datos_tiendas: dict[str, dict] = {}
    ws_dat = wb["DATOS"]
    for row in ws_dat.iter_rows(min_row=2, values_only=True):
        tienda = row[1]  # col B
        if tienda is None or str(tienda).strip() == "":
            continue
        cc  = row[2]  # col C
        doc = row[3]  # col D
        datos_tiendas[str(tienda).upper().strip()] = {
            "cc":  int(cc)  if cc  is not None else 0,
            "doc": int(doc) if doc is not None else 0,
        }

    return invenarios, datos_tiendas


def _cargar_con_gspread() -> tuple[dict, dict]:
    """Lee catálogo desde Google Sheets usando service account."""
    import gspread
    import json as _json
    from google.oauth2.service_account import Credentials

    SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
    SPREADSHEET_ID = "1JzKIDiMmjqVD-iYXTAjxk4wqPvdassNMZJjsNP_UtQI"

    creds_json = os.getenv("GOOGLE_SHEETS_CREDENTIALS")
    if creds_json:
        creds = Credentials.from_service_account_info(_json.loads(creds_json), scopes=SCOPES)
    else:
        creds = Credentials.from_service_account_file("credentials.json", scopes=SCOPES)

    gc = gspread.authorize(creds)
    sp = gc.open_by_key(SPREADSHEET_ID)

    invenarios: dict[str, dict] = {}
    for row in sp.worksheet("INVENARIOS").get_all_values()[2:]:
        ref = row[1] if len(row) > 1 else ""
        if not ref or not str(ref).strip():
            continue
        producto = row[2] if len(row) > 2 else ""
        cta_inv  = row[4] if len(row) > 4 else ""
        invenarios[str(ref).strip()] = {
            "producto": str(producto).strip(),
            "cta_inv":  str(cta_inv).strip(),
        }

    datos_tiendas: dict[str, dict] = {}
    for row in sp.worksheet("DATOS").get_all_values()[1:]:
        tienda = row[1] if len(row) > 1 else ""
        if not tienda or not str(tienda).strip():
            continue
        try:
            cc = int(row[2]) if len(row) > 2 and row[2] else 0
        except Exception:
            cc = 0
        try:
            doc = int(row[3]) if len(row) > 3 and row[3] else 0
        except Exception:
            doc = 0
        datos_tiendas[str(tienda).upper().strip()] = {"cc": cc, "doc": doc}

    return invenarios, datos_tiendas


def _cargar_con_fuente() -> tuple[dict, dict, str]:
    """Intenta Google Sheets primero; si falla, usa archivo local como fallback."""
    try:
        inv, tiendas = _cargar_con_gspread()
        return inv, tiendas, "gsheets"
    except Exception:
        pass

    st.warning(
        "⚠️ No se pudo conectar al catálogo en línea. Se está usando una copia local "
        "que puede estar desactualizada. Haz clic en 'Actualizar catálogo' para reintentar."
    )
    if not EXCEL_PATH.exists():
        return {}, {}, "error"
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, keep_vba=True)
    inv, tiendas = _parse_wb(wb)
    return inv, tiendas, "local"


if "invenarios" not in st.session_state:
    with st.spinner("Cargando referencias de inventarios..."):
        _inv, _tiendas, _fuente = _cargar_con_fuente()
    st.session_state["invenarios"]    = _inv
    st.session_state["datos_tiendas"] = _tiendas
    st.session_state["excel_fuente"]  = _fuente

invenarios    = st.session_state["invenarios"]
datos_tiendas = {_normalizar_ciudad(k): v for k, v in st.session_state["datos_tiendas"].items()}
_fuente       = st.session_state.get("excel_fuente", "local")

if not invenarios:
    st.error(
        "❌ No se pudo cargar el Excel desde OneDrive ni localmente. "
        "Agrega CARGUE FACTURAS DE COMPRA F3 ACT.xlsm a la raíz del proyecto."
    )
    st.stop()

# ─── TÍTULO Y CABECERA (Dashboard Top Bar) ──────────────────────────────────────

st.title("ContaFlow Yamaha")
st.markdown("##### Motor Inteligente de Generación PRN 🚀")

# Contenedor de Estado del Sistema y Herramientas (En lugar del Sidebar)
with st.container():
    c1, c2, c3, c4 = st.columns([1.5, 1.5, 1.5, 1])
    
    with c1:
        if _fuente == "gsheets":
            st.success("🟢 Sincronizado (Google Sheets)")
        else:
            st.warning("🟠 Modo Offline (Local)")
    with c2:
        st.info(f"📦 {len(invenarios):,} referencias en memoria")
    with c3:
        if api_key:
            st.success("🤖 Core IA: En línea")
        else:
            st.error("🔴 Error: Sin conexión IA")
    with c4:
        if st.button("🔄 Refrescar DB", type="secondary", use_container_width=True):
            for k in ["invenarios", "excel_fuente"]:
                st.session_state.pop(k, None)
            st.rerun()

st.divider()

# ─── PASO 1: SUBIR FACTURAS (Dropzone) ────────────────────────────────────────

st.markdown("### 📥 Paso 1: Carga de Facturas CPFE")
st.caption("Arrastra aquí los archivos PDF de Incolmotos Yamaha. Procesaremos todos en lote automáticamente.")

uploaded_files = st.file_uploader(
    "Arrastra tus facturas en formato PDF aquí",
    type=["pdf"],
    accept_multiple_files=True,
    label_visibility="collapsed"
)

if "facturas_procesadas" not in st.session_state:
    st.session_state.facturas_procesadas = {}
if "procesando" not in st.session_state:
    st.session_state.procesando = False
def iniciar_proceso():
    st.session_state.procesando = True

if uploaded_files:
    st.markdown("<br>", unsafe_allow_html=True)
    # Avisar si alguna factura ya fue procesada en esta sesión
    ya_procesadas = [f.name for f in uploaded_files if f.name in st.session_state.facturas_procesadas]
    for nombre_ya in ya_procesadas:
        with st.container():
            col_w, col_b = st.columns([4, 1])
            with col_w:
                st.info(f"💡 **{nombre_ya}** fue procesada recientemente. El sistema cargará su resultado guardado para ahorrar tiempo.")
            with col_b:
                if st.button("🗑️ Vaciar memoria", key=f"reprocess_{nombre_ya}", use_container_width=True, help="Fuerza a la IA a leerla de nuevo."):
                    del st.session_state.facturas_procesadas[nombre_ya]
                    st.rerun()

    st.button(
        "🚀 Iniciar Procesamiento Inteligente",
        type="primary",
        on_click=iniciar_proceso,
        disabled=st.session_state.procesando,
        use_container_width=True,
    )


def _extraer_con_pdfplumber(pdf_bytes: bytes, nombre: str):
    """
    Extractor determinístico de ítems usando pdfplumber.
    No usa IA. Retorna mismo formato que _extraer_iterativa, o None si falla/descuadra.
    """
    import pdfplumber, io, re

    # Artefactos de layout de Incolmotos que aparecen superpuestos en ciertas líneas
    ARTIFACT_RE = re.compile(r'^(?:\d{6}-EFPC|\d{2}/\d{2}/\d{2})\s+')
    ITEM_RE = re.compile(
        r'^(\d+)\s+'            # número de ítem (grupo 1 — clave dedup)
        r'(\S+)\s+'             # referencia (grupo 2)
        r'(.+?)\s+'             # descripcion (grupo 3)
        r'UNIDAD\s+'            # marcador obligatorio
        r'([\d\.]+)\s+'         # cantidad (grupo 4)
        r'[\d,\.]+\s+'          # precio_unitario (ignorado)
        r'(?:[\d,\.]+\s+)?'     # descuento % (opcional, ignorado)
        r'([\d,\.]+)$'          # valor_total (grupo 5)
    )
    SKIP_RE = re.compile(r'^\d+\s+(Pedido|Pag[i]?na|Son:|Pronto|Cant\.)')

    try:
        items = []
        seen_item_nums = set()
        header_text = ''

        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                words = page.extract_words(x_tolerance=3)
                rows = {}
                for w in words:
                    y = int(w['top'])   # int() tolera ±0.1px entre label y valor numérico
                    if y not in rows:
                        rows[y] = []
                    rows[y].append((w['x0'], w['text']))
                for y in sorted(rows.keys()):
                    line = ' '.join(t for _, t in sorted(rows[y]))
                    header_text += line + '\n'
                    clean = ARTIFACT_RE.sub('', line).strip()
                    if not clean or not clean[0].isdigit():
                        continue
                    if SKIP_RE.match(clean):
                        continue
                    m = ITEM_RE.match(clean)
                    if not m:
                        continue
                    item_num = m.group(1)
                    if item_num in seen_item_nums:
                        continue
                    seen_item_nums.add(item_num)
                    items.append({
                        'referencia': m.group(2).strip(),
                        'descripcion': m.group(3).strip(),
                        'cantidad': float(m.group(4)),
                        'valor_total': float(m.group(5).replace(',', '')),
                        'tiene_iva': True,
                    })

        if not items:
            return None

        # Header fields
        m = re.search(r'(CPFE-\d+)', header_text)
        numero_factura = m.group(1) if m else ''

        m = re.search(r'(\d{4}-\d{2}-\d{2})\s+\d{2}:\d{2}:\d{2}', header_text)
        fecha = m.group(1) if m else ''

        m = re.search(r'Ciudad:\s+([A-ZÁÉÍÓÚÑ]+)', header_text)
        ciudad = m.group(1) if m else ''

        m = re.search(r'Direcci[oó]n:\s+([^\n]+)', header_text)
        direccion_entrega = m.group(1).strip() if m else ''

        m = re.search(r'Bruto[\s\n]+([\d,\.]+)', header_text)
        subtotal = float(m.group(1).replace(',', '')) if m else 0.0

        m = re.search(r'\bIVA[\s\n]+([\d,\.]+)', header_text)
        iva_total = float(m.group(1).replace(',', '')) if m else 0.0

        m = re.search(r'HASTA EL (\d{2}-\d{2}-\d{4})', header_text)
        if m:
            d, mo, y_str = m.group(1).split('-')
            fecha_vto = f'{y_str}{mo}{d}'
        else:
            m = re.search(r'Vencimiento:\s+(\d{4}-\d{2}-\d{2})', header_text)
            fecha_vto = m.group(1).replace('-', '') if m else '00000000'

        # Solo retornar si la suma cuadra con el subtotal
        suma = sum(i['valor_total'] for i in items)
        if subtotal > 0 and abs(suma - subtotal) > 10.0:
            return None  # Descuadre — dejar que Claude maneje este PDF

        return {
            'header': {
                'numero_factura': numero_factura,
                'fecha': fecha,
                'ciudad': ciudad,
                'direccion_entrega': direccion_entrega,
                'subtotal': subtotal,
                'iva_total': iva_total,
                'fecha_vto': fecha_vto,
            },
            'items': items,
        }

    except Exception:
        return None  # Cualquier fallo → fallback a Claude


def _extraer_iterativa(client, pdf_b64, nombre, base_prompt):
    todos_items = []
    seen_keys = set()
    header = None
    iteracion = 0
    hay_mas = True
    
    while hay_mas and iteracion < 50:
        iteracion += 1
        
        if iteracion == 1:
            prompt = f"{base_prompt}\n\nMUESTRA LOS PRIMEROS 50 ÍTEMS. Si hay más de 50 ítems, marca 'hay_mas_items: true'."
        else:
            ultimos = todos_items[-3:] if len(todos_items) >= 3 else todos_items
            ultimos_str = json.dumps(ultimos, ensure_ascii=False)
            ultima_ref = todos_items[-1].get('referencia', '') if todos_items else ''
            subtotal_esp = float(header.get('subtotal', 0) or 0) if header else 0
            suma_act = sum(float(item.get('valor_total', 0) or 0) for item in todos_items)
            aviso_faltantes = (
                f"ATENCIÓN: El subtotal de la factura es ${subtotal_esp:,.2f} pero solo he extraído "
                f"${suma_act:,.2f}. Aún faltan ítems en el PDF — revisa TODAS las páginas.\n\n"
                if subtotal_esp > 0 and (subtotal_esp - suma_act) > 10.0 else ""
            )
            prompt = (
                f"Extrae SOLO el JSON válido. Usa el MISMO exacto formato JSON requerido anteriormente.\n"
                f"Misma factura. Ya extraje {len(todos_items)} ítems. Los últimos 3 fueron:\n"
                f"{ultimos_str}\n\n"
                f"{aviso_faltantes}"
                f"Extrae los siguientes 50 ítems que aparecen DESPUÉS de la referencia \"{ultima_ref}\".\n"
                f"NUNCA repitas ítems ya extraídos. NUNCA fusiones repetidos.\n"
                f"Si terminaste envía los ítems y pon \"hay_mas_items\": false. Si siguen más pon true."
            )
        
        try:
            resp = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=8192,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_b64}},
                            {"type": "text", "text": prompt}
                        ]
                    },
                    {
                        "role": "assistant",
                        "content": "{"
                    }
                ]
            )
            # Como forzamos que inicie con '{', lo concatenamos al inicio de lo que el modelo escupa
            raw = "{" + resp.content[0].text
            
            # Limpiar posible basura final (ej. si el modelo cierra con markdown)
            raw = raw.strip()
            if raw.endswith("```"):
                raw = raw[:-3].strip()
                
            data = json.loads(raw)
        except json.JSONDecodeError as je:
            st.error(f"Error procesando página {iteracion} de {nombre}: {je}")
            st.error(f"**TEXTO ORIGINAL DEVUELTO POR LA IA:**\n\n{raw}")
            return None
        except Exception as e:
            st.error(f"Error procesando API en página {iteracion} de {nombre}: {e}")
            return None
        
        if iteracion == 1:
            # We extract header fields from the root level as per the new flat prompt format
            header = {
                "numero_factura": data.get("numero_factura", ""),
                "fecha": data.get("fecha", ""),
                "ciudad": data.get("ciudad", ""),
                "direccion_entrega": data.get("direccion_entrega", ""),
                "subtotal": data.get("subtotal", 0.0),
                "iva_total": data.get("iva_total", 0.0),
                "fecha_vto": data.get("fecha_vto", "")
            }
        
        nuevos = data.get('items', [])

        # Dedup: filtrar items que la IA repitio entre iteraciones
        # Clave incluye descripcion para no eliminar items legitimos con misma ref/precio
        nuevos_unicos = []
        for item in nuevos:
            key = (
                str(item.get('referencia', '')).strip(),
                str(item.get('descripcion', '')).strip()[:40],
                item.get('cantidad', 0),
                item.get('valor_total', 0)
            )
            if key not in seen_keys:
                nuevos_unicos.append(item)
                seen_keys.add(key)

        if len(nuevos_unicos) < len(nuevos):
            st.info(f"🔁 {nombre}: Se descartaron {len(nuevos) - len(nuevos_unicos)} items duplicados en iteracion {iteracion}.")

        todos_items.extend(nuevos_unicos)
        hay_mas = data.get('hay_mas_items', False)

        # Recovery: Haiku dijo 'done' pero la suma no cuadra con el subtotal del PDF
        if not hay_mas and header and todos_items:
            subtotal_esperado = float(header.get('subtotal', 0) or 0)
            suma_actual = sum(float(item.get('valor_total', 0) or 0) for item in todos_items)
            if subtotal_esperado > 0 and (subtotal_esperado - suma_actual) > 10.0:
                hay_mas = True
                st.info(
                    f"🔍 {nombre}: Haiku reportó fin pero hay ítems faltantes. "
                    f"Subtotal PDF=${subtotal_esperado:,.0f} | Extraído=${suma_actual:,.0f} | "
                    f"Faltan≈${subtotal_esperado - suma_actual:,.0f}. Continuando extracción..."
                )

    if hay_mas:
        import streamlit as st
        st.warning(f"⚠️ {nombre}: Factura muy grande. Se procesaron {len(todos_items)} ítems. Verifica que el total cuadre.")
    
    if not header:
        header = {}
    return {"header": header, "items": todos_items}

def _clean_float(val):
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        v = val.replace("$", "").replace(",", "").strip()
        try:
            return float(v)
        except:
            return 0.0
    return 0.0

if st.session_state.procesando and uploaded_files:
    facturas_extraidas = []
    facturas_fallidas = []

    with st.spinner("Procesando facturas, por favor espera..."):
        try:
            client = anthropic.Anthropic(api_key=api_key)

            prompt = (
                "<instructions>\n"
                "Eres un extractor experto de datos de facturas de Incolmotos Colombia.\n"
                "Tu única tarea es extraer datos del documento PDF adjunto y estructurarlos exactamente bajo el esquema JSON proveido.\n"
                "</instructions>\n\n"
                "<anti_hallucination_rules>\n"
                "1. Extrae ÚNICAMENTE información que esté literalmente presente en el documento.\n"
                "2. No infieras, no 'limpies', ni inventes ceros ni caracteres extra. Transcribe el texto EXACTAMENTE como se ve.\n"
                "3. Las referencias de Yamaha generalmente tienen 10 o 12 caracteres. Extráelas literal, sin rellenar con ceros fantasmas.\n"
                "4. La factura puede tener múltiples páginas. Extrae ABSOLUTAMENTE TODOS los ítems de TODAS las páginas.\n"
                "5. Crea exactamente un objeto por cada número de ítem visible en el PDF. NUNCA fusiones ítems similares o repetidos.\n"
                "</anti_hallucination_rules>\n\n"
                "<schema_json>\n"
                "Debes responder SOLO y EXCLUSIVAMENTE emitiendo código JSON puramente válido, sin formato markdown, sin intro ni outro.\n"
                "{\n"
                "  \"numero_factura\": \"CPFE-XXXXXX (número completo con prefijo)\",\n"
                "  \"fecha\": \"YYYY-MM-DD\",\n"
                "  \"ciudad\": \"solo la primera palabra del campo Ciudad del encabezado, ejemplo: si dice NEIVA HUILA devuelve NEIVA, si dice GIRARDOT CUNDINAMARCA devuelve GIRARDOT\",\n"
                "  \"direccion_entrega\": \"dirección de envío o sucursal destino tal como aparece en el PDF (ej. CR 5 20 39)\",\n"
                "  \"subtotal\": 0.0,\n"
                "  \"iva_total\": 0.0,\n"
                "  \"fecha_vto\": \"YYYYMMDD — buscar primero el texto 'HASTA EL DD-MM-YYYY' (fecha pronto pago, convertir de DD-MM-YYYY a YYYYMMDD); si no existe, usar el campo Vencimiento del encabezado (ya en formato YYYYMMDD)\",\n"
                "  \"items\": [\n"
                "    {\n"
                "      \"referencia\": \"literal de la columna Referencia, sin agregar caracteres ni ceros extra\",\n"
                "      \"descripcion\": \"columna Producto, texto completo sin truncar\",\n"
                "      \"cantidad\": 1,\n"
                "      \"valor_total\": 0.0,\n"
                "      \"tiene_iva\": true\n"
                "    }\n"
                "  ]\n"
                "}\n"
                "</schema_json>"
            )

            bar     = st.progress(0)
            status  = st.empty()
            total   = len(uploaded_files)

            for idx, archivo in enumerate(uploaded_files, 1):
                status.markdown(f"🧠 Procesando **{idx}/{total}**: `{archivo.name}`...")

                # Cache: si ya fue procesada en esta sesión, reusar resultado
                if archivo.name in st.session_state.facturas_procesadas:
                    facturas_extraidas.append(st.session_state.facturas_procesadas[archivo.name])
                    bar.progress(idx / total)
                    continue

                try:
                    pdf_bytes = archivo.read()

                    # Intento 1: extractor determinístico (pdfplumber, sin IA)
                    extracted = _extraer_con_pdfplumber(pdf_bytes, archivo.name)
                    if extracted:
                        st.info(f"⚡ {archivo.name}: Extracción determinística OK ({len(extracted['items'])} ítems).")
                    else:
                        # Intento 2: Claude Haiku (fallback para PDFs no estándar)
                        pdf_b64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")
                        extracted = _extraer_iterativa(client, pdf_b64, archivo.name, prompt)

                    if not extracted:
                        facturas_fallidas.append(archivo.name)
                        continue
                    
                    datos = extracted["header"]
                    datos["items"] = extracted["items"]
                    
                    # Validar numero_factura mayusculas
                    if datos.get("numero_factura"):
                        datos["numero_factura"] = str(datos["numero_factura"]).upper()
                        
                    # Clean floats
                    datos["subtotal"] = _clean_float(datos.get("subtotal", 0.0))
                    datos["iva_total"] = _clean_float(datos.get("iva_total", 0.0))
                    for item in datos["items"]:
                        item["valor_total"] = _clean_float(item.get("valor_total", 0.0))
                        
                    # Validate math (Subtotal vs Sum of items) — BLOQUEAR si no cuadra
                    suma_items = sum(item["valor_total"] for item in datos["items"])
                    if abs(suma_items - datos["subtotal"]) > 10.0:
                        st.error(
                            f"🚫 {archivo.name}: DESCUADRE DETECTADO.\n"
                            f"Suma de Items = ${suma_items:,.2f} | Subtotal PDF = ${datos['subtotal']:,.2f}\n"
                            f"Diferencia: ${abs(suma_items - datos['subtotal']):,.2f}\n\n"
                            f"El PRN NO se generara hasta corregir esto. Re-sube la factura."
                        )
                        facturas_fallidas.append(archivo.name)
                        continue
                    
                    # Validation of fecha_vto
                    fvto = str(datos.get("fecha_vto", "")).strip()
                    if fvto and fvto.isdigit() and len(fvto) == 8:
                        mes = int(fvto[4:6])
                        dia = int(fvto[6:8])
                        if not (1 <= mes <= 12 and 1 <= dia <= 31):
                            datos["fecha_vto"] = "00000000"
                    else:
                        datos["fecha_vto"] = "00000000"
                    
                    datos["_nombre_archivo"] = archivo.name
                    st.session_state.facturas_procesadas[archivo.name] = datos
                    facturas_extraidas.append(datos)

                except json.JSONDecodeError:
                    facturas_fallidas.append(archivo.name)
                    st.error(
                        f"Error procesando {archivo.name}: la IA no devolvió "
                        "JSON válido. Intenta de nuevo."
                    )
                except Exception as e:
                    facturas_fallidas.append(archivo.name)
                    st.error(f"Error procesando {archivo.name}: {e}")

                bar.progress(idx / total)

        finally:
            status.empty()
            bar.empty()
            st.session_state["facturas_extraidas"] = facturas_extraidas
            st.session_state.procesando = False

    if facturas_fallidas:
        st.error(f"❌ Fallaron {len(facturas_fallidas)} facturas: {', '.join(facturas_fallidas)}. Por favor reinténtalas.")

    if facturas_extraidas:
        st.success(f"✅ {len(facturas_extraidas)} factura(s) procesada(s) correctamente")

# ─── PASO 3: VALIDACIÓN ───────────────────────────────────────────────────────

facturas = st.session_state.get("facturas_extraidas", [])

if not facturas:
    st.stop()

st.divider()
st.markdown("### 📊 Paso 2: Validación Visual")

refs_faltantes = []
for fac in facturas:
    for item in fac.get("items", []):
        ref = str(item.get("referencia", "")).strip()
        
        # [AUTO-CORRECCIÓN DE CEROS FANTASMAS (AI/OCR hallucinations)]
        if ref and ref not in invenarios:
            # A veces la IA agrega ceros de más al final ("5VV1764159000" en vez de "5VV176415900")
            if ref.endswith("00") and ref[:-2] in invenarios:
                ref = ref[:-2]
                item["referencia"] = ref
            elif ref.endswith("0") and ref[:-1] in invenarios:
                ref = ref[:-1]
                item["referencia"] = ref
            # A veces la IA omite ceros al final
            elif (ref + "0") in invenarios:
                ref = ref + "0"
                item["referencia"] = ref
            elif (ref + "00") in invenarios:
                ref = ref + "00"
                item["referencia"] = ref

        if ref not in invenarios:
            refs_faltantes.append({
                "Factura":     fac.get("numero_factura", "?"),
                "Referencia":  ref,
                "Descripción": item.get("descripcion", ""),
            })

if refs_faltantes:
    import pandas as pd
    refs_unicas: dict[str, str] = {}
    for r in refs_faltantes:
        if r["Referencia"] not in refs_unicas:
            refs_unicas[r["Referencia"]] = r.get("Descripción", "")

    n = len(refs_unicas)
    st.error(f"⚠️ Se detectaron **{n} referencia(s)** sin código contable asignado en memoria.", icon="🚨")
    
    with st.expander("Ver lista de items en conflicto cruzado", expanded=False):
        st.dataframe(pd.DataFrame(refs_faltantes), use_container_width=True, hide_index=True)

    st.markdown("##### 🛠️ Clasificador Manual:")
    codigos_ingresados = {}
    with st.container(border=True):
        for ref, desc in refs_unicas.items():
            col1, col2, col3 = st.columns([2, 1.5, 1.5])
            with col1:
                st.markdown(f"**Referencia:** `{ref}`  \n<small>{desc}</small>", unsafe_allow_html=True)
            with col2:
                codigo = st.text_input("Ingresa Cód. Producto Siigo", key=f"cod_{ref}", placeholder="Ej: 0020089...")
            with col3:
                if codigo:
                    cta = _calcular_cta_inv(codigo)
                    st.success(f"**CTA:** `{cta}`")
                    codigos_ingresados[ref] = {"producto": codigo, "descripcion": desc, "cta_inv": cta}
                else:
                    st.caption("Esperando...")

    todos_completos = len(codigos_ingresados) == len(refs_unicas)

    if todos_completos:
        if st.button("✅ Confirmar y Guardar en Nube", use_container_width=True, type="primary"):
            exitos = 0
            for ref, datos in codigos_ingresados.items():
                if _guardar_referencia_en_sheets(ref, datos["producto"], datos["descripcion"], datos["cta_inv"]):
                    exitos += 1
            if exitos == len(codigos_ingresados):
                st.success(f"✅ {exitos} referencias inyectadas al cerebro maestro.")
                for k in ["invenarios", "excel_fuente"]:
                    st.session_state.pop(k, None)
                st.rerun()
            else:
                st.error("Fallo de red al persistir en nube.")
    else:
        faltantes_count = len(refs_unicas) - len(codigos_ingresados)
        st.warning(f"💡 Debes codificar {faltantes_count} ítem(s) más para continuar.")

    st.stop()

# ── Todas las referencias existen — mostrar resumen por factura ────────────────
import pandas as pd
st.success("Toda la facturación cuadró perfectamente con el catálogo contable.", icon="✅")

tab_labels = [f"📄 {fac.get('numero_factura', '?')}" for fac in facturas]
tabs = st.tabs(tab_labels)

for fac, tab in zip(facturas, tabs):
    with tab:
        num = fac.get("numero_factura", "?")
        subtotal  = float(fac.get("subtotal",  0))
        iva_total = float(fac.get("iva_total", 0))
        total_fac = subtotal + iva_total

        m1, m2, m3 = st.columns(3)
        m1.metric("Subtotal P.",  f"${subtotal:,.0f}")
        m2.metric("Impuestos (IVA)", f"${iva_total:,.0f}")
        m3.metric("Coste TOTAL", f"${total_fac:,.0f}")
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        filas = []
        for item in fac.get("items", []):
            ref  = str(item.get("referencia", "")).strip()
            look = invenarios.get(ref, {})
            filas.append({
                "Referencia":  ref,
                "Descripción": item.get("descripcion", ""),
                "Cuenta":      look.get("cta_inv", "—"),
                "Cantidad":    item.get("cantidad", 0),
                "Valor Total": item.get("valor_total", 0),
            })
        st.dataframe(pd.DataFrame(filas), use_container_width=True, hide_index=True)

# ── Validar códigos producto duplicados dentro de cada factura ────────────────

prods_duplicados = []
for fac in facturas:
    vistos: dict[str, str] = {}  # producto -> primera referencia que lo usó
    for item in fac.get("items", []):
        ref  = str(item.get("referencia", "")).strip()
        prod = invenarios.get(ref, {}).get("producto", "")
        if not prod:
            continue
        if prod in vistos and vistos[prod] != ref:
            prods_duplicados.append({
                "Factura":          fac.get("numero_factura", "?"),
                "Referencia":       ref,
                "Descripción":      item.get("descripcion", ""),
                "Código Producto":  prod,
                "Conflicto con":    vistos[prod],
            })
        else:
            vistos[prod] = ref

if prods_duplicados:
    st.error(
        "⛔ Error: referencias con código contable duplicado. "
        "Verificar en Siigo antes de continuar."
    )
    st.dataframe(
        pd.DataFrame(prods_duplicados),
        use_container_width=True,
        hide_index=True,
    )
    st.stop()

# ─── PASO 4 + 5: GENERACIÓN Y DESCARGA PRN ───────────────────────────────────

def generar_prn_lines(
    factura_data: dict,
    inv: dict,
    tiendas: dict,
) -> list[str]:
    ciudad = _normalizar_ciudad(factura_data["ciudad"])
    
    if "IBAGU" in ciudad:
        direccion_entrega = factura_data.get("direccion_entrega", "")
        cc, doc = _resolver_tienda_ibague(direccion_entrega)
    else:
        if ciudad not in tiendas:
            raise ValueError(f"Ciudad '{ciudad}' no encontrada en Excel DATOS.")
        cc  = tiendas[ciudad]["cc"]
        doc = tiendas[ciudad]["doc"]
        


    num_doc = (
        factura_data["numero_factura"]
        .replace("CPFE-", "")
        .replace("CPFE",  "")
        .strip()
    )
    fecha     = factura_data["fecha"].replace("-", "")      # YYYYMMDD
    _fvto     = str(factura_data.get("fecha_vto", "")).strip()
    fecha_vto = _fvto if (_fvto.isdigit() and len(_fvto) == 8) else "00000000"

    def fmt_line(
        sec: int,
        cuenta: str,
        producto: str,
        descripcion: str,
        deb_cred: str,
        valor: float,
        cantidad: float,
    ) -> str:
        import unicodedata
        descripcion_limpia = unicodedata.normalize('NFKD', str(descripcion)).encode('ascii', 'ignore').decode('utf-8').upper()
        line = (
            "P"                                          # TIPO           1
            + str(doc).zfill(3)                          # COD.COMP       3
            + num_doc.zfill(11)                          # NUM.DOC       11
            + str(sec).zfill(5)                          # SEC            5
            + NIT_INCOLMOTOS.zfill(13)                   # NIT           13
            + "000"                                      # SUCURSAL       3
            + str(cuenta).ljust(10)[:10]                 # CUENTA        10
            + str(producto).ljust(13)[:13]               # PRODUCTO      13
            + fecha                                      # FECHA DOC      8
            + str(cc).zfill(4)                           # C.COSTO        4
            + "000"                                      # S.COSTO        3
            + str(descripcion_limpia).ljust(50)[:50]     # DESCRIPCION   50
            + deb_cred                                   # DEB.CRED       1
            + str(int(round(valor * 100))).zfill(15)     # VR.MOVIM      15
            + "000000000000000"                          # BASE RET      15
            + "0000"                                     # COD.VEND       4
            + "0000"                                     # COD.CIUDAD     4
            + "000"                                      # COD.ZONA       3
            + str(cc).zfill(4)                           # COD.BODEGA     4
            + "000"                                      # COD.UB         3
            + str(int(round(cantidad * 100000))).zfill(15)  # CANTIDAD   15
            + " "                                        # TIPO D.CRUCE   1
            + "000"                                      # COD.D.CRUCE    3
            + "00000000000"                              # NUM.D.CRUCE   11
            + "000"                                      # SEC.DOC.CRUCE  3
            + "00000000"                                 # FECHA VENC     8
            + "0000"                                     # COD.FORMA PAGO 4
            + "00"                                       # COD.BANCO      2
        )
        if len(line) != 220:
            raise ValueError(f"Error interno generando PRN: línea tiene {len(line)} caracteres (esperado 220). Contacta al administrador.")
        return line

    lines: list[str] = []
    sec = 1

    # Líneas D — ítems de inventario
    for item in factura_data.get("items", []):
        ref  = str(item["referencia"]).strip()
        look = inv[ref]
        if not look.get("cta_inv", "").strip():
            raise ValueError(f"La referencia '{ref}' no tiene cuenta contable en el catálogo. Verifica el Excel.")
        lines.append(fmt_line(
            sec         = sec,
            cuenta      = look["cta_inv"],
            producto    = look["producto"],
            descripcion = item.get("descripcion", ""),
            deb_cred    = "D",
            valor       = float(item.get("valor_total", 0)),
            cantidad    = float(item.get("cantidad", 1)),
        ))
        sec += 1

    # Línea C — CXP proveedor (campos de cruce y vencimiento específicos)
    total_fac = float(factura_data.get("subtotal", 0)) + float(factura_data.get("iva_total", 0))
    cxp_line = (
        "P"                                              # TIPO           1
        + str(doc).zfill(3)                              # COD.COMP       3
        + num_doc.zfill(11)                              # NUM.DOC       11
        + str(sec).zfill(5)                              # SEC            5
        + NIT_INCOLMOTOS.zfill(13)                       # NIT           13
        + "000"                                          # SUCURSAL       3
        + "2205010000"                                   # CUENTA        10
        + "0000000000000"                                # PRODUCTO      13
        + fecha                                          # FECHA DOC      8
        + str(cc).zfill(4)                               # C.COSTO        4
        + "000"                                          # S.COSTO        3
        + "INCOLMOTOS SAS".ljust(50)[:50]                # DESCRIPCION   50
        + "C"                                            # DEB.CRED       1
        + str(int(round(total_fac * 100))).zfill(15)     # VR.MOVIM      15
        + "000000000000000"                              # BASE RET      15
        + "0000"                                         # COD.VEND       4
        + "0000"                                         # COD.CIUDAD     4
        + "000"                                          # COD.ZONA       3
        + str(cc).zfill(4)                               # COD.BODEGA     4
        + "000"                                          # COD.UB         3
        + "000000000000000"                              # CANTIDAD      15
        + "P"                                            # TIPO D.CRUCE   1
        + str(doc).zfill(3)                              # COD.D.CRUCE    3
        + num_doc.zfill(11)                              # NUM.D.CRUCE   11
        + "001"                                          # SEC.DOC.CRUCE  3
        + fecha_vto                                      # FECHA VENC     8
        + "0000"                                         # COD.FORMA PAGO 4
        + "00"                                           # COD.BANCO      2
    )
    if len(cxp_line) != 220:
        raise ValueError(f"Error interno generando PRN: línea tiene {len(cxp_line)} caracteres (esperado 220). Contacta al administrador.")
    lines.append(cxp_line)
    sec += 1

    # Línea D — IVA descontable (si aplica)
    iva = float(factura_data.get("iva_total", 0))
    if iva > 0:
        lines.append(fmt_line(
            sec         = sec,
            cuenta      = "2408020100",
            producto    = "0000000000000",
            descripcion = "INCOLMOTOS SAS",
            deb_cred    = "D",
            valor       = iva,
            cantidad    = 1,
        ))

    return lines


st.divider()
st.markdown("### 🏁 Paso 3: Exportación a Siigo")
st.caption("Asegúrate de haber validado visualmente los montos arriba antes de generar los enlaces.")

if st.button("🚀 Empaquetar y Generar Archivos PRN", type="primary", use_container_width=True):
    archivos_prn: list[tuple[str, bytes]] = []
    error_generacion = False

    for fac in facturas:
        num = (fac.get("numero_factura", "0").replace("CPFE-", "").replace("CPFE",  "").strip())
        nombre = f"INT{num}.prn"

        try:
            lines   = generar_prn_lines(fac, invenarios, datos_tiendas)
            content = "\r\n".join(lines) + "\r\n"
            archivos_prn.append((nombre, content.encode("latin-1", errors="replace")))
        except ValueError as e:
            st.error(str(e))
            error_generacion = True
            break
        except Exception as e:
            st.error(f"Error generando {nombre}: {e}")
            error_generacion = True
            break

    if not error_generacion and archivos_prn:
        st.balloons()
        st.success("🎉 ¡Conversión Exitosa! Tus archivos están listos para descargarse.")
        
        with st.container(border=True):
            if len(archivos_prn) == 1:
                nombre, datos_prn = archivos_prn[0]
                st.download_button(
                    label              = f"⬇️ Descargar contabilidad de {nombre}",
                    data               = datos_prn,
                    file_name          = nombre,
                    mime               = "application/octet-stream",
                    use_container_width=True,
                )
            else:
                import io
                import zipfile
                zip_buf = io.BytesIO()
                with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
                    for nombre, datos_prn in archivos_prn:
                        zf.writestr(nombre, datos_prn)
                zip_buf.seek(0)
                st.download_button(
                    label              = "📦 ⬇️ Descargar Paquete ZIP (Múltiples Facturas)",
                    data               = zip_buf,
                    file_name          = "Lote_ContaFlow.zip",
                    mime               = "application/zip",
                    use_container_width=True,
                )

        # ── Tabla de verificación cruzada al final ──────────────────────────────────────────
        st.markdown("<br>##### 🔎 Auditoría Cruzada Final", unsafe_allow_html=True)
        verificacion = []
        for fac in facturas:
            subtotal  = float(fac.get("subtotal",  0))
            iva_total = float(fac.get("iva_total", 0))
            verificacion.append({
                "Nº de Factura":  fac.get("numero_factura", "?"),
                "Items": len(fac.get("items", [])),
                "Subtotal": f"${subtotal:,.0f}",
                "IVA":      f"${iva_total:,.0f}",
                "T. A PAGAR":    f"${subtotal + iva_total:,.0f}",
            })
        st.dataframe(pd.DataFrame(verificacion), use_container_width=True, hide_index=True)
        st.caption("✔️ Verifica siempre que el T. A PAGAR coincida con el final del archivo PDF impreso.")
