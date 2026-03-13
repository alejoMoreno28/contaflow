"""
test_yamaha.py — Prueba end-to-end de la lógica de yamaha_app.py
Sin Streamlit: simula exactamente lo que haría un usuario.
"""

import base64
import json
import os
import re
import sys
from pathlib import Path

import anthropic
import openpyxl
from dotenv import load_dotenv

load_dotenv()

EXCEL_PATH     = Path("CARGUE FACTURAS DE COMPRA F3 ACT.xlsm")
PDF_PATH       = Path("cpfe-750304.pdf")
NIT_INCOLMOTOS = "890916911"

# ─── PASO 1: CARGAR EXCEL ─────────────────────────────────────────────────────

print("=" * 60)
print("PASO 1: Cargando Excel de inventarios...")
print("=" * 60)

if not EXCEL_PATH.exists():
    print(f"ERROR: No se encontró {EXCEL_PATH}")
    sys.exit(1)

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, keep_vba=True)

invenarios: dict[str, dict] = {}
ws_inv = wb["INVENARIOS"]
for row in ws_inv.iter_rows(min_row=3, values_only=True):
    ref = row[1]
    if ref is None or str(ref).strip() == "":
        continue
    producto = row[2]
    cta_inv  = row[4]
    invenarios[str(ref).strip()] = {
        "producto": str(producto).strip() if producto is not None else "",
        "cta_inv":  str(cta_inv).strip()  if cta_inv  is not None else "",
    }

datos_tiendas: dict[str, dict] = {}
ws_dat = wb["DATOS"]
for row in ws_dat.iter_rows(min_row=2, values_only=True):
    tienda = row[1]
    if tienda is None or str(tienda).strip() == "":
        continue
    cc  = row[2]
    doc = row[3]
    datos_tiendas[str(tienda).upper().strip()] = {
        "cc":  int(cc)  if cc  is not None else 0,
        "doc": int(doc) if doc is not None else 0,
    }

print(f"  ✅ {len(invenarios)} referencias en INVENARIOS")
print(f"  ✅ {len(datos_tiendas)} tiendas en DATOS: {list(datos_tiendas.keys())}")

# ─── PASO 2: PROCESAR PDF CON HAIKU ──────────────────────────────────────────

print()
print("=" * 60)
print("PASO 2: Procesando PDF con Claude Haiku...")
print("=" * 60)

api_key = os.environ.get("ANTHROPIC_API_KEY", "")
if not api_key:
    print("ERROR: ANTHROPIC_API_KEY no configurada")
    sys.exit(1)

if not PDF_PATH.exists():
    print(f"ERROR: No se encontró {PDF_PATH}")
    sys.exit(1)

pdf_bytes = PDF_PATH.read_bytes()
pdf_b64   = base64.standard_b64encode(pdf_bytes).decode("utf-8")
print(f"  PDF: {PDF_PATH.name} ({len(pdf_bytes):,} bytes)")

prompt = (
    "Eres un extractor de datos de facturas de Incolmotos Colombia. "
    "Extrae exactamente estos campos y responde SOLO con JSON válido, "
    "sin explicaciones, sin markdown, sin bloques de código:\n"
    "{\n"
    "  \"numero_factura\": \"CPFE-XXXXXX (número completo con prefijo)\",\n"
    "  \"fecha\": \"YYYY-MM-DD\",\n"
    "  \"ciudad\": \"solo la primera palabra del campo Ciudad del encabezado, "
    "ejemplo: si dice NEIVA HUILA devuelve NEIVA, "
    "si dice GIRARDOT CUNDINAMARCA devuelve GIRARDOT\",\n"
    "  \"subtotal\": 0.0,\n"
    "  \"iva_total\": 0.0,\n"
    "  \"items\": [\n"
    "    {\n"
    "      \"referencia\": \"exactamente como aparece en columna Referencia, sin espacios extra\",\n"
    "      \"descripcion\": \"columna Producto, máximo 50 caracteres\",\n"
    "      \"cantidad\": 1,\n"
    "      \"valor_total\": 0.0,\n"
    "      \"tiene_iva\": true\n"
    "    }\n"
    "  ]\n"
    "}"
)

client = anthropic.Anthropic(api_key=api_key)
message = client.messages.create(
    model="claude-haiku-4-5-20251001",
    max_tokens=4096,
    messages=[
        {
            "role": "user",
            "content": [
                {
                    "type": "document",
                    "source": {
                        "type":       "base64",
                        "media_type": "application/pdf",
                        "data":       pdf_b64,
                    },
                },
                {"type": "text", "text": prompt},
            ],
        }
    ],
)

raw = message.content[0].text.strip()
raw = re.sub(r"^```(?:json)?\s*\n?", "", raw)
raw = re.sub(r"\n?```\s*$", "", raw).strip()

try:
    factura = json.loads(raw)
except json.JSONDecodeError as e:
    print(f"ERROR: JSON inválido de Haiku: {e}")
    print("Respuesta raw:")
    print(raw)
    sys.exit(1)

print(f"  ✅ Factura extraída: {factura.get('numero_factura')}")
print(f"     Fecha:    {factura.get('fecha')}")
print(f"     Ciudad:   {factura.get('ciudad')}")
print(f"     Subtotal: ${factura.get('subtotal', 0):,.0f}")
print(f"     IVA:      ${factura.get('iva_total', 0):,.0f}")
print(f"     Ítems:    {len(factura.get('items', []))}")

# ─── PASO 3: VALIDAR REFERENCIAS ─────────────────────────────────────────────

print()
print("=" * 60)
print("PASO 3: Validando referencias contra INVENARIOS...")
print("=" * 60)

faltantes = []
for item in factura.get("items", []):
    ref = str(item.get("referencia", "")).strip()
    if ref in invenarios:
        look = invenarios[ref]
        print(f"  ✅ {ref} → cuenta {look['cta_inv']} | producto {look['producto'][:30]}")
    else:
        print(f"  ❌ {ref} → NO ENCONTRADA en INVENARIOS")
        faltantes.append(ref)

if faltantes:
    print(f"\nERROR: {len(faltantes)} referencia(s) no encontrada(s): {faltantes}")
    sys.exit(1)

print(f"\n  ✅ Todas las referencias encontradas")

# ─── PASO 4: GENERAR PRN ─────────────────────────────────────────────────────

print()
print("=" * 60)
print("PASO 4: Generando archivo PRN...")
print("=" * 60)

ciudad = factura["ciudad"].upper().strip().split()[0]
if ciudad not in datos_tiendas:
    print(f"ERROR: Ciudad '{ciudad}' no en DATOS. Disponibles: {list(datos_tiendas.keys())}")
    sys.exit(1)

cc  = datos_tiendas[ciudad]["cc"]
doc = datos_tiendas[ciudad]["doc"]
print(f"  Ciudad: {ciudad} → CC={cc}, DOC={doc}")

num_doc = (
    factura["numero_factura"]
    .replace("CPFE-", "")
    .replace("CPFE",  "")
    .strip()
)
fecha = factura["fecha"].replace("-", "")

def fmt_line(sec, cuenta, producto, descripcion, deb_cred, valor, cantidad):
    line = (
        "P"
        + str(doc).zfill(3)
        + num_doc.zfill(11)
        + str(sec).zfill(5)
        + NIT_INCOLMOTOS.zfill(13)
        + "000"
        + str(cuenta).ljust(10)[:10]
        + str(producto).ljust(13)[:13]
        + fecha
        + str(cc).zfill(4)
        + "000"
        + str(descripcion).ljust(50)[:50]
        + deb_cred
        + str(int(round(valor * 100))).zfill(15)
        + "000000000000000"
        + "0000"
        + "0000"
        + "000"
        + str(cc).zfill(4)
        + "000"
        + str(int(round(cantidad * 100000))).zfill(15)
        + " "
        + "000"
        + "00000000000"
        + "000"
        + "00000000"
        + "0000"
        + "00"
    )
    assert len(line) == 220, f"Línea tiene {len(line)} chars (esperado 220)"
    return line

lines = []
sec = 1

for item in factura.get("items", []):
    ref  = str(item["referencia"]).strip()
    look = invenarios[ref]
    lines.append(fmt_line(
        sec         = sec,
        cuenta      = look["cta_inv"],
        producto    = look["producto"],
        descripcion = item.get("descripcion", ""),
        deb_cred    = "D",
        valor       = float(item.get("valor_total", 0)),
        cantidad    = float(item.get("cantidad", 1)),
    ))
    sec += 1

total_fac = float(factura.get("subtotal", 0)) + float(factura.get("iva_total", 0))
lines.append(fmt_line(
    sec=sec, cuenta="2205010000", producto="0000000000000",
    descripcion="INCOLMOTOS SAS", deb_cred="C",
    valor=total_fac, cantidad=1,
))
sec += 1

iva = float(factura.get("iva_total", 0))
if iva > 0:
    lines.append(fmt_line(
        sec=sec, cuenta="2408020100", producto="0000000000000",
        descripcion="INCOLMOTOS SAS", deb_cred="D",
        valor=iva, cantidad=1,
    ))

# ─── PASO 5: VERIFICAR 220 CHARS ─────────────────────────────────────────────

print()
print("=" * 60)
print("PASO 5: Verificando longitud de cada línea...")
print("=" * 60)

all_ok = True
for i, line in enumerate(lines, 1):
    if len(line) != 220:
        print(f"  ❌ Línea {i}: {len(line)} chars")
        all_ok = False
    else:
        print(f"  ✅ Línea {i}: 220 chars OK")

if not all_ok:
    sys.exit(1)

# ─── GUARDAR PRN ──────────────────────────────────────────────────────────────

nombre_prn = f"INT{num_doc}.prn"
content    = "\r\n".join(lines) + "\r\n"
Path(nombre_prn).write_bytes(content.encode("latin-1"))
print(f"\n  ✅ Archivo guardado: {nombre_prn} ({len(lines)} líneas)")

# ─── PRIMERAS 3 LÍNEAS ────────────────────────────────────────────────────────

print()
print("=" * 60)
print("Primeras 3 líneas del PRN:")
print("=" * 60)
for i, line in enumerate(lines[:3], 1):
    print(f"\n  [Línea {i}]")
    print(f"  TIPO       : {line[0]}")
    print(f"  COD.COMP   : {line[1:4]}")
    print(f"  NUM.DOC    : {line[4:15]}")
    print(f"  SEC        : {line[15:20]}")
    print(f"  NIT        : {line[20:33]}")
    print(f"  SUCURSAL   : {line[33:36]}")
    print(f"  CUENTA     : {line[36:46]}")
    print(f"  PRODUCTO   : {line[46:59]}")
    print(f"  FECHA      : {line[59:67]}")
    print(f"  C.COSTO    : {line[67:71]}")
    print(f"  S.COSTO    : {line[71:74]}")
    print(f"  DESCRIPCION: {line[74:124]}")
    print(f"  DEB.CRED   : {line[124]}")
    print(f"  VR.MOVIM   : {line[125:140]}")
    print(f"  Raw        : {line}")

print()
print("=" * 60)
print("✅ PRUEBA END-TO-END COMPLETADA EXITOSAMENTE")
print(f"   Archivo generado: {nombre_prn}")
print("=" * 60)
