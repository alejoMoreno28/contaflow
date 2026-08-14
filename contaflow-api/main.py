from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Dict, Any, Optional
import base64
import json
import os
import unicodedata
from pathlib import Path
import zipfile
import io
import re
import anthropic
import openpyxl
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv

from yamaha_catalog import build_catalog_row, parse_inventory_rows
from yamaha_prn import (
    PrnValidationError,
    build_accounting_plan,
    generar_prn_lines,
)
from yamaha_rules import YamahaRuleError, calcular_cta_inv

load_dotenv()

app = FastAPI(title="ContaFlow API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

EXCEL_PATH = Path(__file__).parent.parent / "CARGUE FACTURAS DE COMPRA F3 ACT.xlsm"
NIT_INCOLMOTOS = "890916911"

# --- Globals caching ---
CACHE_INVENARIOS = {}
CACHE_TIENDAS = {}

# --- Funciones Utilitarias (Migradas de Streamlit) ---

def _resolver_tienda_ibague(direccion: str):
    d = unicodedata.normalize('NFD', str(direccion)).encode('ascii','ignore').decode().upper()
    PRINCIPAL_KW = ['CR 5','CRR 5','CARRERA 5','CRA 5','QUINTA','20 39','20-39','B 1 CARMEN','PRINCIPAL']
    SEXTA_KW    = ['CR 6','CRR 6','CARRERA 6','CRA 6','BELALCAZAR','BRR BELALCAZAR','SEXTA','25 40','25-40']
    
    for kw in PRINCIPAL_KW:
        if kw in d:
            return 7, 14
    for kw in SEXTA_KW:
        if kw in d:
            return 1, 1
    return 1, 1

def _normalizar_ciudad(texto: str) -> str:
    return (
        unicodedata.normalize("NFD", texto)
        .encode("ascii", "ignore")
        .decode("utf-8")
        .upper()
        .strip()
        .split()[0]
    )

def _calcular_cta_inv(codigo_producto: str) -> str:
    try:
        return calcular_cta_inv(codigo_producto)
    except YamahaRuleError:
        return ""

def _cargar_con_gspread():
    import gspread
    from google.oauth2.service_account import Credentials
    
    SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
    SPREADSHEET_ID = "1JzKIDiMmjqVD-iYXTAjxk4wqPvdassNMZJjsNP_UtQI"

    creds_json = os.environ.get("GOOGLE_SHEETS_CREDENTIALS")
    if creds_json:
        creds = Credentials.from_service_account_info(json.loads(creds_json), scopes=SCOPES)
    else:
        creds = Credentials.from_service_account_file(Path(__file__).parent.parent / "credentials.json", scopes=SCOPES)

    gc = gspread.authorize(creds)
    sp = gc.open_by_key(SPREADSHEET_ID)

    inv = parse_inventory_rows(
        sp.worksheet("INVENARIOS").get_all_values()[2:]
    )
    dt = {}

    for row in sp.worksheet("DATOS").get_all_values()[1:]:
        tienda = row[1] if len(row) > 1 else ""
        if not tienda or not str(tienda).strip(): continue
        try: cc = int(row[2]) if len(row) > 2 and row[2] else 0
        except Exception: cc = 0
        try: doc = int(row[3]) if len(row) > 3 and row[3] else 0
        except Exception: doc = 0
        dt[str(tienda).upper().strip()] = {"cc": cc, "doc": doc}

    return inv, dt

def _parse_wb(wb):
    inv = parse_inventory_rows(
        wb["INVENARIOS"].iter_rows(min_row=3, values_only=True)
    )
    dt = {}
    for row in wb["DATOS"].iter_rows(min_row=2, values_only=True):
        if row[1] and str(row[1]).strip():
            dt[str(row[1]).upper().strip()] = {
                "cc": int(row[2]) if row[2] else 0,
                "doc": int(row[3]) if row[3] else 0,
            }
    return inv, dt

def get_catalogo():
    global CACHE_INVENARIOS, CACHE_TIENDAS
    if not CACHE_INVENARIOS:
        try:
            inv, tiendas = _cargar_con_gspread()
            CACHE_INVENARIOS, CACHE_TIENDAS = inv, tiendas
        except Exception as e:
            print("Fallo GSheets, intentando local", e)
            if EXCEL_PATH.exists():
                wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, keep_vba=True)
                inv, tiendas = _parse_wb(wb)
                CACHE_INVENARIOS, CACHE_TIENDAS = inv, tiendas
    
    return CACHE_INVENARIOS, {_normalizar_ciudad(k): v for k, v in CACHE_TIENDAS.items()}


def _extraer_iterativa(client, pdf_b64, base_prompt):
    todos_items = []
    header = None
    iteracion = 0
    hay_mas = True
    
    while hay_mas and iteracion < 50:
        iteracion += 1
        if iteracion == 1:
            prompt = f"{base_prompt}\n\nMUESTRA LOS PRIMEROS 50 ÍTEMS. Si hay más de 50 ítems, marca 'hay_mas_items: true'."
        else:
            ultimos = todos_items[-3:] if len(todos_items) >= 3 else todos_items
            ultimos_str = json.dumps(ultimos, ensure_ascii=False)
            ultima_ref = todos_items[-1].get('referencia', '') if todos_items else ''
            prompt = f"Extrae SOLO el JSON válido. Misma factura. Los últimos 3 fueron: {ultimos_str}. Extrae los siguientes 50 ítems después de \"{ultima_ref}\". Si terminaste envía los ítems y pon \"hay_mas_items\": false. Si siguen más pon true."
            
        try:
            resp = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=8192,
                messages=[
                    {"role": "user", "content": [
                        {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_b64}},
                        {"type": "text", "text": prompt}
                    ]},
                    {"role": "assistant", "content": "{"}
                ]
            )
            raw = "{" + resp.content[0].text
            raw = raw.strip()
            if raw.endswith("```"): raw = raw[:-3].strip()
            data = json.loads(raw)
        except Exception as e:
            print("Error parsing AI JSON", e)
            return None
        
        if iteracion == 1:
            header = {
                "numero_factura": data.get("numero_factura", ""),
                "fecha": data.get("fecha", ""),
                "ciudad": data.get("ciudad", ""),
                "direccion_entrega": data.get("direccion_entrega", ""),
                "subtotal": data.get("subtotal", 0.0),
                "iva_total": data.get("iva_total", 0.0),
                "fecha_vto": data.get("fecha_vto", "")
            }
        
        todos_items.extend(data.get('items', []))
        hay_mas = data.get('hay_mas_items', False)
        
    return {"header": header or {}, "items": todos_items}

def _clean_float(val):
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        v = val.replace("$", "").replace(",", "").strip()
        try:
            return float(v)
        except:
            return 0.0
    return 0.0

# --- ENDPOINTS ---

@app.get("/health")
def health_check():
    return {"status": "ok"}

@app.post("/api/extract")
async def extract_invoice(file: UploadFile = File(...)):
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        raise HTTPException(status_code=500, detail="API Key de Anthropic no configurada")
    
    client = anthropic.Anthropic(api_key=api_key)
    inv, tiendas = get_catalogo()
    
    pdf_bytes = await file.read()
    pdf_b64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")
    
    prompt = (
        "<instructions>\n"
        "Eres un extractor experto de datos de facturas de Incolmotos Colombia.\n"
        "Tu única tarea es extraer datos del documento PDF adjunto y estructurarlos exactamente bajo el esquema JSON proveido.\n"
        "</instructions>\n\n"
        "<anti_hallucination_rules>\n"
        "1. Extrae ÚNICAMENTE información que esté literalmente presente en el documento.\n"
        "2. No infieras, no 'limpies', ni inventes ceros ni caracteres extra. Transcribe el texto EXACTAMENTE como se ve.\n"
        "3. Las referencias de Yamaha generalmente tienen 10 o 12 caracteres. Extráelas literal, sin rellenar con ceros fantasmas.\n"
        "4. La factura puede tener múltiples páginas. Extrae ABSOLUTAMENTE TODOS los ítems de TODAS las páginas.\n"
        "5. Crea exactamente un objeto por cada número de ítem visible en el PDF. NUNCA fusiones ítems similares o repetidos.\n"
        "</anti_hallucination_rules>\n\n"
        "<schema_json>\n"
        "Debes responder SOLO y EXCLUSIVAMENTE emitiendo código JSON puramente válido, sin formato markdown, sin intro ni outro.\n"
        "{\n"
        "  \"numero_factura\": \"CPFE-XXXXXX\",\n"
        "  \"fecha\": \"YYYY-MM-DD\",\n"
        "  \"ciudad\": \"solo la primera palabra del campo Ciudad del encabezado\",\n"
        "  \"direccion_entrega\": \"dirección de envío o sucursal destino\",\n"
        "  \"subtotal\": 0.0,\n"
        "  \"iva_total\": 0.0,\n"
        "  \"fecha_vto\": \"YYYYMMDD\",\n"
        "  \"items\": [\n"
        "    {\n"
        "      \"referencia\": \"literal de la columna Referencia\",\n"
        "      \"descripcion\": \"columna Producto\",\n"
        "      \"cantidad\": 1,\n"
        "      \"valor_total\": 0.0,\n"
        "      \"tiene_iva\": true\n"
        "    }\n"
        "  ]\n"
        "}\n"
        "</schema_json>"
    )
    
    extracted = _extraer_iterativa(client, pdf_b64, prompt)
    if not extracted:
        raise HTTPException(status_code=500, detail="Fallo al extraer JSON de la IA (posible timeout)")
        
    datos = extracted["header"]
    datos["items"] = extracted["items"]
    
    if datos.get("numero_factura"):
        datos["numero_factura"] = str(datos["numero_factura"]).upper()
        
    datos["subtotal"] = _clean_float(datos.get("subtotal", 0.0))
    datos["iva_total"] = _clean_float(datos.get("iva_total", 0.0))
    for item in datos["items"]:
        item["valor_total"] = _clean_float(item.get("valor_total", 0.0))
        
    fvto = str(datos.get("fecha_vto", "")).strip()
    if fvto and fvto.isdigit() and len(fvto) == 8:
        mes = int(fvto[4:6])
        dia = int(fvto[6:8])
        if not (1 <= mes <= 12 and 1 <= dia <= 31):
            datos["fecha_vto"] = "00000000"
    else:
        datos["fecha_vto"] = "00000000"
        
    suma_items = sum(item["valor_total"] for item in datos["items"])
    descuadre_math = False
    if abs(suma_items - datos["subtotal"]) > 10.0:
        descuadre_math = True
    
    missing_refs = []
    
    for item in datos.get("items", []):
        ref = str(item.get("referencia", "")).strip()
        # Autocorrección de ceros fantasmas
        if ref and ref not in inv:
            if ref.endswith("00") and ref[:-2] in inv:
                ref = ref[:-2]
                item["referencia"] = ref
            elif ref.endswith("0") and ref[:-1] in inv:
                ref = ref[:-1]
                item["referencia"] = ref
            elif (ref + "0") in inv:
                ref = ref + "0"
                item["referencia"] = ref
            elif (ref + "00") in inv:
                ref = ref + "00"
                item["referencia"] = ref
                
        if ref not in inv:
            missing_refs.append({
                "referencia": ref,
                "descripcion": item.get("descripcion", ""),
            })

    # Extraer items duplicados en la factura como en streamlit
    prods_duplicados = []
    vistos = {}
    for item in datos.get("items", []):
        ref = str(item.get("referencia", "")).strip()
        prod = inv.get(ref, {}).get("producto", "")
        if not prod: continue
        if prod in vistos and vistos[prod] != ref:
            prods_duplicados.append({
                "referencia": ref,
                "descripcion": item.get("descripcion", ""),
                "codigo_producto": prod,
            })
        else:
            vistos[prod] = ref

    accounting_preview = None
    accounting_error = None
    if not missing_refs and not prods_duplicados and not descuadre_math:
        try:
            plan = build_accounting_plan(datos, inv)
            accounting_preview = {
                "capitalized_vat": float(plan.capitalized_vat),
                "deductible_vat": float(plan.deductible_vat),
                "total_payable": float(plan.credit_total),
                "items": [
                    {
                        "referencia": movement.reference,
                        "cuenta": movement.account,
                        "tratamiento_iva": movement.treatment,
                        "base": float(movement.base),
                        "iva_mayor_costo": float(movement.capitalized_vat),
                        "costo_inventario": float(movement.amount),
                    }
                    for movement in plan.item_movements
                ],
            }
        except PrnValidationError as exc:
            accounting_error = str(exc)

    return {
        "status": "success",
        "factura": datos,
        "missing_refs": missing_refs,
        "duplicados": prods_duplicados,
        "descuadre_math": descuadre_math,
        "accounting_preview": accounting_preview,
        "accounting_error": accounting_error,
    }

class RefData(BaseModel):
    referencia: str
    producto: str
    descripcion: str
    tratamiento_iva: Optional[str] = ""
    cta_inv: Optional[str] = None


def _open_inventory_worksheet():
    import gspread
    from google.oauth2.service_account import Credentials

    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    spreadsheet_id = "1JzKIDiMmjqVD-iYXTAjxk4wqPvdassNMZJjsNP_UtQI"
    creds_json = os.environ.get("GOOGLE_SHEETS_CREDENTIALS")
    if creds_json:
        creds = Credentials.from_service_account_info(
            json.loads(creds_json),
            scopes=scopes,
        )
    else:
        creds = Credentials.from_service_account_file(
            Path(__file__).parent.parent / "credentials.json",
            scopes=scopes,
        )
    return gspread.authorize(creds).open_by_key(spreadsheet_id).worksheet("INVENARIOS")

@app.post("/api/add-reference")
def add_reference(refs: List[RefData]):
    from datetime import date
    fecha = date.today().strftime("%Y-%m-%d")
    try:
        rows = [
            build_catalog_row(
                ref.referencia,
                ref.producto,
                ref.descripcion,
                ref.tratamiento_iva,
                creation_date=fecha,
            )
            for ref in refs
        ]
    except YamahaRuleError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    ws = _open_inventory_worksheet()
    existing_refs = {
        str(row[1]).strip()
        for row in ws.get_all_values()[2:]
        if len(row) > 1 and str(row[1]).strip()
    }
    duplicates = [row[1] for row in rows if row[1] in existing_refs]
    if duplicates:
        raise HTTPException(
            status_code=409,
            detail=f"Referencias ya existentes: {', '.join(duplicates)}",
        )

    ws.append_rows(rows, value_input_option="USER_ENTERED")

    global CACHE_INVENARIOS
    for row in rows:
        CACHE_INVENARIOS[row[1]] = {
            "producto": row[2].lstrip("'"),
            "descripcion": row[3],
            "cta_inv": row[4],
            "linea": row[6],
            "grupo": row[7],
            "tratamiento_iva": row[8],
        }

    return {
        "status": "success",
        "references": [
            {
                "referencia": row[1],
                "producto": row[2].lstrip("'"),
                "cta_inv": row[4],
                "tratamiento_iva": row[8],
            }
            for row in rows
        ],
    }

class FacturaRequest(BaseModel):
    facturas: List[Dict[str, Any]]

@app.post("/api/generate-prn")
def generate_prn(data: FacturaRequest):
    inv, tiendas = get_catalogo()
    archivos_prn = []
    for fac in data.facturas:
        try:
            lines = generar_prn_lines(fac, inv, tiendas)
        except PrnValidationError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        num_doc = (
            str(fac.get("numero_factura", ""))
            .replace("CPFE-", "")
            .replace("CPFE", "")
            .strip()
        )
        content = "\r\n".join(lines) + "\r\n"
        archivos_prn.append((f"INT{num_doc}.prn", content.encode("latin-1", errors="replace")))

    if len(archivos_prn) == 1:
        headers = {"Content-Disposition": f"attachment; filename={archivos_prn[0][0]}"}
        return StreamingResponse(io.BytesIO(archivos_prn[0][1]), headers=headers, media_type="application/octet-stream")
    else:
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            for nombre, datos_prn in archivos_prn:
                zf.writestr(nombre, datos_prn)
        zip_buf.seek(0)
        headers = {"Content-Disposition": "attachment; filename=Lote_ContaFlow.zip"}
        return StreamingResponse(zip_buf, headers=headers, media_type="application/zip")
