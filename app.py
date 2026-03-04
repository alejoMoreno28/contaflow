"""
app.py — ContaFlow (Streamlit)
Wizard lineal: Paso 1 (Subir + Conectar) → Paso 2 (Revisar) → Paso 3 (Sincronizar + Excel)
"""

import json
import os
import re
import tempfile
from datetime import datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from extractor import process_pdf, process_image, IMAGE_MEDIA_TYPES

try:
    from alegra_client import AlegraClient, AlegraAuthError, AlegraAPIError, friendly_error
    from memory_manager import get_item_memory, save_invoice_memory, normalize_item_key
    _ALEGRA_OK = True
except ImportError:
    _ALEGRA_OK = False

try:
    from ai_suggester import suggest_account_for_item
except ImportError:
    def suggest_account_for_item(*a, **kw):  # noqa: E302
        return None

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# AUTENTICACIÓN
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

_VALID_EMAIL    = "demo@contaflow.co"
_VALID_PASSWORD = "ContaFlow2024"

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PAGE CONFIG
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

st.set_page_config(
    page_title="ContaFlow",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={"Get Help": None, "Report a bug": None, "About": None},
)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CSS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

PROFESSIONAL_CSS = """
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

    * { font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif !important; }

    :root {
        --blue:        #0066FF;
        --blue-dark:   #0052CC;
        --blue-light:  #E8F0FF;
        --blue-mid:    #4D94FF;
        --green:       #10B981;
        --green-light: #ECFDF5;
        --text:        #0F172A;
        --text-soft:   #475569;
        --text-xsoft:  #94A3B8;
        --bg:          #FFFFFF;
        --bg-soft:     #F8FAFF;
        --border:      #E2E8F0;
        --shadow-sm:   0 1px 3px rgba(0,0,0,0.08);
        --shadow-md:   0 4px 16px rgba(0,102,255,0.10);
        --radius:      14px;
        --radius-sm:   8px;
        --radius-pill: 999px;
    }

    html, body, [data-testid="stAppViewContainer"], [data-testid="stBaseViewContainer"] {
        background: var(--bg) !important;
        color: var(--text) !important;
    }

    [data-testid="stToolbar"],
    [data-testid="stMainMenu"],
    footer { display: none !important; }

    [data-testid="stAppViewContainer"]::before {
        content: "";
        position: fixed;
        top: 0; left: 0; right: 0;
        height: 4px;
        background: linear-gradient(90deg, var(--blue) 0%, var(--blue-mid) 100%);
        z-index: 999;
        pointer-events: none;
    }

    [data-testid="stMainBlockContainer"] {
        padding: 0 24px !important;
        max-width: 1200px !important;
        margin: 0 auto !important;
    }

    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #F8FAFF 0%, #FFFFFF 100%) !important;
        border-right: 1px solid var(--border) !important;
    }
    [data-testid="stSidebarContent"] { padding: 32px 24px !important; }

    button[kind="primary"] {
        background: linear-gradient(135deg, var(--blue) 0%, var(--blue-dark) 100%) !important;
        color: #fff !important;
        border: none !important;
        border-radius: var(--radius-pill) !important;
        padding: 14px 28px !important;
        font-weight: 700 !important;
        font-size: 1rem !important;
        box-shadow: 0 4px 16px rgba(0,102,255,0.35) !important;
        transition: all 0.22s cubic-bezier(0.4,0,0.2,1) !important;
        letter-spacing: -0.01em !important;
    }
    button[kind="primary"]:hover {
        background: linear-gradient(135deg, var(--blue-dark) 0%, #003D99 100%) !important;
        box-shadow: 0 8px 28px rgba(0,102,255,0.45) !important;
        transform: translateY(-2px) !important;
    }

    button[kind="secondary"] {
        background: transparent !important;
        border: 2px solid var(--border) !important;
        color: var(--text) !important;
        border-radius: var(--radius-pill) !important;
        padding: 12px 24px !important;
        font-weight: 600 !important;
        transition: all 0.22s cubic-bezier(0.4,0,0.2,1) !important;
    }
    button[kind="secondary"]:hover {
        border-color: var(--blue) !important;
        color: var(--blue) !important;
        background: var(--blue-light) !important;
    }

    .download-btn button {
        background: linear-gradient(135deg, var(--green) 0%, #059669 100%) !important;
        color: #fff !important;
        border: none !important;
        border-radius: var(--radius-pill) !important;
        padding: 16px 32px !important;
        font-weight: 700 !important;
        font-size: 1.0625rem !important;
        box-shadow: 0 4px 16px rgba(16,185,129,0.35) !important;
        transition: all 0.22s cubic-bezier(0.4,0,0.2,1) !important;
    }
    .download-btn button:hover {
        background: linear-gradient(135deg, #059669 0%, #047857 100%) !important;
        box-shadow: 0 8px 28px rgba(16,185,129,0.45) !important;
        transform: translateY(-2px) !important;
    }

    [data-testid="stFileUploadDropzone"] {
        border: 2px dashed var(--blue) !important;
        border-radius: var(--radius) !important;
        background: var(--blue-light) !important;
        padding: 48px 32px !important;
    }

    h1 {
        font-size: 2.5rem !important;
        font-weight: 800 !important;
        letter-spacing: -0.04em !important;
        line-height: 1.15 !important;
        color: var(--text) !important;
        margin: 32px 0 8px !important;
    }
    h2 {
        font-size: 1.875rem !important;
        font-weight: 800 !important;
        letter-spacing: -0.04em !important;
        color: var(--text) !important;
        margin: 0 !important;
    }
    h3 {
        font-size: 1.125rem !important;
        font-weight: 700 !important;
        color: var(--text) !important;
        margin: 24px 0 12px !important;
    }

    .step-header {
        display: flex;
        align-items: center;
        gap: 14px;
        margin: 36px 0 18px;
    }
    .step-num {
        display: inline-flex;
        align-items: center;
        justify-content: center;
        width: 42px; height: 42px;
        border-radius: 50%;
        background: var(--blue);
        color: #fff;
        font-weight: 800;
        font-size: 1.125rem;
        flex-shrink: 0;
        box-shadow: 0 4px 12px rgba(0,102,255,0.28);
    }

    .bloque-card {
        background: var(--bg-soft);
        border: 1.5px solid var(--border);
        border-radius: var(--radius);
        padding: 20px 22px;
        margin-bottom: 14px;
    }
    .bloque-title {
        font-size: 0.8125rem !important;
        font-weight: 700 !important;
        color: var(--blue) !important;
        text-transform: uppercase;
        letter-spacing: 0.07em;
        margin-bottom: 14px !important;
    }

    [data-testid="stMetricContainer"] {
        background: var(--bg) !important;
        border: 1.5px solid var(--border) !important;
        border-radius: var(--radius) !important;
        padding: 24px !important;
        box-shadow: var(--shadow-sm) !important;
        transition: all 0.22s cubic-bezier(0.4,0,0.2,1) !important;
    }
    [data-testid="stMetricContainer"]:hover {
        border-color: var(--blue) !important;
        box-shadow: var(--shadow-md) !important;
        transform: translateY(-4px) !important;
    }
    [data-testid="stMetricValue"] {
        font-size: 2rem !important;
        font-weight: 800 !important;
        color: var(--blue) !important;
        letter-spacing: -0.04em !important;
    }
    [data-testid="stMetricLabel"] {
        font-size: 0.8125rem !important;
        color: var(--text-soft) !important;
        font-weight: 600 !important;
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }

    [data-testid="stDataFrameContainer"] {
        border: 1.5px solid var(--border) !important;
        border-radius: var(--radius) !important;
        overflow: hidden !important;
        box-shadow: var(--shadow-sm) !important;
    }

    [data-testid="stAlert"] {
        border-radius: var(--radius) !important;
        border-left: 4px solid !important;
        padding: 14px 18px !important;
    }

    hr {
        border: none !important;
        border-top: 1px solid var(--border) !important;
        margin: 32px 0 !important;
    }

    @media (max-width: 900px) {
        h1 { font-size: 1.875rem !important; }
        h2 { font-size: 1.5rem !important; }
    }
</style>
"""

st.markdown(PROFESSIONAL_CSS, unsafe_allow_html=True)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# LOGIN
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    st.markdown("""
    <style>
    html, body,
    [data-testid="stAppViewContainer"],
    [data-testid="stBaseViewContainer"] {
        background: #1a1a2e !important;
    }
    [data-testid="stMainBlockContainer"] {
        background: transparent !important;
        max-width: 100% !important;
        padding: 0 16px !important;
    }
    [data-testid="stSidebar"]  { display: none !important; }
    [data-testid="stHeader"]   { background: transparent !important; }
    [data-testid="stAppViewContainer"]::before { display: none !important; }

    div[data-testid="stHorizontalBlock"] > div:nth-child(2) > div[data-testid="stVerticalBlockBorderWrapper"] > div,
    div[data-testid="stHorizontalBlock"] > div:nth-child(2) > div[data-testid="stVerticalBlock"] {
        background: #ffffff !important;
        border-radius: 20px !important;
        padding: 48px 40px !important;
        box-shadow: 0 24px 80px rgba(0,0,0,0.35) !important;
        margin-top: 60px !important;
    }

    div[data-testid="stHorizontalBlock"] > div:nth-child(2) [data-testid="stImage"] {
        display: flex !important;
        justify-content: center !important;
    }

    div[data-testid="stHorizontalBlock"] > div:nth-child(2) [data-testid="stTextInput"] input {
        border: 1.5px solid #E2E8F0 !important;
        border-radius: 10px !important;
        background: #F8FAFF !important;
        padding: 12px 14px !important;
        font-size: 0.9375rem !important;
        color: #0F172A !important;
    }
    div[data-testid="stHorizontalBlock"] > div:nth-child(2) [data-testid="stTextInput"] input:focus {
        border-color: #0066FF !important;
        box-shadow: 0 0 0 3px rgba(0,102,255,0.15) !important;
        background: #fff !important;
    }

    div[data-testid="stHorizontalBlock"] > div:nth-child(2) [data-testid="stForm"] button {
        background: #0066FF !important;
        background-image: none !important;
        color: #fff !important;
        border: none !important;
        border-radius: 10px !important;
        padding: 14px !important;
        font-size: 1rem !important;
        font-weight: 700 !important;
        box-shadow: 0 4px 16px rgba(0,102,255,0.30) !important;
        transform: none !important;
        letter-spacing: 0 !important;
        margin-top: 8px !important;
    }
    div[data-testid="stHorizontalBlock"] > div:nth-child(2) [data-testid="stForm"] button:hover {
        background: #0052CC !important;
        box-shadow: 0 6px 24px rgba(0,102,255,0.40) !important;
        transform: translateY(-1px) !important;
    }
    </style>
    """, unsafe_allow_html=True)

    _, _mid, _ = st.columns([1, 1.5, 1])
    with _mid:
        _logo = Path(__file__).parent / "logo.png"
        if _logo.exists():
            st.image(str(_logo), width=120)
        else:
            st.markdown(
                '<div style="text-align:center;font-size:3rem;margin-bottom:4px;">⚡</div>',
                unsafe_allow_html=True,
            )

        st.markdown("""
        <div style="text-align:center;margin:14px 0 28px;">
            <div style="font-size:1.5rem;font-weight:800;color:#0066FF;letter-spacing:-0.03em;">
                Bienvenido a ContaFlow
            </div>
            <div style="font-size:0.875rem;color:#64748B;margin-top:6px;">
                Ingresa tus credenciales para continuar
            </div>
        </div>
        """, unsafe_allow_html=True)

        with st.form("login_form"):
            _email    = st.text_input("Correo electrónico", placeholder="usuario@empresa.co")
            _password = st.text_input("Contraseña", type="password", placeholder="••••••••")
            _submitted = st.form_submit_button("Entrar", use_container_width=True, type="primary")

        if _submitted:
            if _email == _VALID_EMAIL and _password == _VALID_PASSWORD:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("Credenciales incorrectas")

    st.stop()

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SESSION STATE — wizard
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

_defaults = {
    "wizard_step":       1,         # 1 | 2 | 3
    "batch_results":     [],        # extracciones crudas de Claude
    "batch_names":       [],        # nombres de archivos
    "reviewed_invoices": [],        # facturas con parametrización (Step 2 → 3)
    "historial":         [],        # facturas confirmadas (para Excel)
    "upload_results":    [],        # resultados subida a Alegra
    # Alegra
    "alegra_connected":  False,
    "alegra_email":      "",
    "alegra_token":      "",
    "alegra_catalogs":   {"accounts": [], "taxes": [], "cost_centers": []},
}
for _k, _v in _defaults.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# HELPERS DE CATÁLOGOS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

_NONE_LABEL = "— (ninguno) —"


def _account_options() -> list[str]:
    accs = st.session_state.alegra_catalogs.get("accounts", [])
    return [_NONE_LABEL] + [f"{a['code']} — {a['name']}" for a in accs]


def _tax_options() -> list[str]:
    taxes = st.session_state.alegra_catalogs.get("taxes", [])
    return [_NONE_LABEL] + [f"{t['name']}" for t in taxes]


def _cc_options() -> list[str]:
    ccs = st.session_state.alegra_catalogs.get("cost_centers", [])
    return [_NONE_LABEL] + [c["name"] for c in ccs]


def _label_to_account_id(label: str) -> int | None:
    if not label or label == _NONE_LABEL:
        return None
    for a in st.session_state.alegra_catalogs.get("accounts", []):
        if f"{a['code']} — {a['name']}" == label:
            return a["id"]
    return None


def _label_to_tax_id(label: str) -> int | None:
    if not label or label == _NONE_LABEL:
        return None
    for t in st.session_state.alegra_catalogs.get("taxes", []):
        if t["name"] == label:
            return t["id"]
    return None


def _label_to_cc_id(label: str) -> int | None:
    if not label or label == _NONE_LABEL:
        return None
    for c in st.session_state.alegra_catalogs.get("cost_centers", []):
        if c["name"] == label:
            return c["id"]
    return None


def _account_id_to_label(account_id: int | None) -> str:
    if account_id is None:
        return _NONE_LABEL
    for a in st.session_state.alegra_catalogs.get("accounts", []):
        if a["id"] == account_id:
            return f"{a['code']} — {a['name']}"
    return _NONE_LABEL


def _tax_id_to_label(tax_id: int | None) -> str:
    if tax_id is None:
        return _NONE_LABEL
    for t in st.session_state.alegra_catalogs.get("taxes", []):
        if t["id"] == tax_id:
            return t["name"]
    return _NONE_LABEL


def _cc_id_to_label(cc_id: int | None) -> str:
    if cc_id is None:
        return _NONE_LABEL
    for c in st.session_state.alegra_catalogs.get("cost_centers", []):
        if c["id"] == cc_id:
            return c["name"]
    return _NONE_LABEL


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# GENERADOR DE EXCEL (intacto)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def crear_excel(historial: list[dict]) -> BytesIO:
    from openpyxl.styles import Alignment, Font, PatternFill

    COLS_RESUMEN = [
        "proveedor",        "nit_proveedor",    "numero_factura",
        "fecha_emision",    "fecha_vencimiento",
        "direccion",        "telefono",
        "comprador",        "nit_comprador",
        "total_bruto",      "subtotal",          "pct_iva",
        "valor_iva",        "total_a_pagar",
        "tipo_comprobante", "centro_costo",      "forma_pago",
        "aplica_rf25",      "rf25_valor",
        "aplica_rf35",      "rf35_valor",
        "aplica_riva",      "riva_valor",
        "aplica_rica",      "tasa_rica",         "rica_valor",
    ]
    COLS_MONEDA = {
        "total_bruto", "subtotal", "valor_iva", "total_a_pagar",
        "rf25_valor",  "rf35_valor", "riva_valor", "rica_valor",
    }

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        fill_blue  = PatternFill(start_color="0066FF", end_color="0066FF", fill_type="solid")
        font_white = Font(color="FFFFFF", bold=True, size=11)
        center     = Alignment(horizontal="center", vertical="center")

        # HOJA 1 — RESUMEN
        df_res  = pd.DataFrame(historial)
        cols_ok = [c for c in COLS_RESUMEN if c in df_res.columns]
        df_res[cols_ok].to_excel(writer, sheet_name="Resumen", index=False)

        ws = writer.sheets["Resumen"]
        for cell in ws[1]:
            cell.fill      = fill_blue
            cell.font      = font_white
            cell.alignment = center

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                col_name = ws.cell(row=1, column=cell.column).value
                if col_name in COLS_MONEDA:
                    cell.number_format = '$ #,##0'

        for col in ws.columns:
            ancho = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(ancho + 4, 35)

        # HOJA 2 — ITEMS
        filas_items = []
        for fac in historial:
            for item in fac.get("items_detalle", []):
                filas_items.append({
                    "# Factura":   fac.get("numero_factura", ""),
                    "Proveedor":   fac.get("proveedor", ""),
                    "Descripción": item.get("Descripción", ""),
                    "Tipo":        item.get("Tipo", ""),
                    "Cuenta PUC":  item.get("Cuenta PUC", ""),
                    "Cantidad":    item.get("Cantidad", ""),
                    "Valor Total": item.get("Valor Total", ""),
                })

        df_items = pd.DataFrame(
            filas_items if filas_items else [],
            columns=["# Factura", "Proveedor", "Descripción",
                     "Tipo", "Cuenta PUC", "Cantidad", "Valor Total"],
        )
        df_items.to_excel(writer, sheet_name="Items", index=False)

        ws2 = writer.sheets["Items"]
        for cell in ws2[1]:
            cell.fill      = fill_blue
            cell.font      = font_white
            cell.alignment = center

        header_items = {cell.value: cell.column for cell in ws2[1]}
        col_vt = header_items.get("Valor Total")
        if col_vt:
            for row in ws2.iter_rows(min_row=2, min_col=col_vt, max_col=col_vt):
                for cell in row:
                    cell.number_format = '$ #,##0'

        for col in ws2.columns:
            ancho = max((len(str(c.value or "")) for c in col), default=10)
            ws2.column_dimensions[col[0].column_letter].width = min(ancho + 4, 45)

    buf.seek(0)
    return buf


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SIDEBAR
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

with st.sidebar:
    st.markdown("""
    <div style="display:flex;align-items:center;gap:12px;margin-bottom:28px;">
        <div style="width:44px;height:44px;background:linear-gradient(135deg,#0066FF,#0052CC);
                    border-radius:12px;display:flex;align-items:center;justify-content:center;
                    box-shadow:0 4px 12px rgba(0,102,255,0.3);">
            <span style="font-size:22px;">⚡</span>
        </div>
        <div>
            <div style="font-size:1.25rem;font-weight:800;color:#0F172A;letter-spacing:-0.03em;">
                Conta<span style="color:#0066FF;">Flow</span>
            </div>
            <div style="font-size:0.75rem;color:#94A3B8;font-weight:600;">Colombia 🇨🇴</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.divider()
    st.markdown("### 🔧 Estado")
    if not os.getenv("ANTHROPIC_API_KEY"):
        st.error("⚠️ ANTHROPIC_API_KEY no configurada")
    else:
        st.success("✅ API Claude activa")

    if st.session_state.alegra_connected:
        st.success(f"✅ Alegra: {st.session_state.alegra_email}")
    else:
        st.info("🔗 Alegra: no conectado")

    st.divider()
    st.markdown("### 📊 Sesión")
    step_labels = {1: "1 — Subir", 2: "2 — Revisar", 3: "3 — Sincronizar"}
    st.caption(f"Paso actual: **{step_labels.get(st.session_state.wizard_step, '?')}**")
    st.metric("Facturas en lote",     len(st.session_state.batch_results))
    st.metric("Facturas en historial", len(st.session_state.historial))

    st.divider()
    if st.button("🔄 Nuevo lote (reiniciar)", use_container_width=True):
        for k in ("wizard_step", "batch_results", "batch_names",
                  "reviewed_invoices", "upload_results"):
            st.session_state[k] = _defaults[k]
        st.rerun()

    st.divider()
    st.markdown("👤 **Demo ContaFlow**")
    if st.button("Cerrar sesión", use_container_width=True):
        st.session_state.authenticated = False
        st.rerun()
    st.caption("**ContaFlow © 2026** • Colombia 🇨🇴")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# HEADER + INDICADOR DE PASOS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

st.markdown("""
<div style="margin-bottom:24px;padding-top:16px;">
    <h1>⚡ ContaFlow</h1>
    <p style="font-size:1.0625rem;color:#475569;margin:0;line-height:1.6;">
        Sube facturas → Extrae con IA → Revisa → Exporta a Excel / Alegra
    </p>
</div>
""", unsafe_allow_html=True)

_ws = st.session_state.wizard_step
_steps = [
    ("1", "Subir y Conectar"),
    ("2", "Revisar Facturas"),
    ("3", "Sincronizar"),
]

_step_html = '<div style="display:flex;gap:0;margin-bottom:32px;border:1.5px solid #E2E8F0;border-radius:12px;overflow:hidden;">'
for _sn, _sl in _steps:
    _active = int(_sn) == _ws
    _done   = int(_sn) < _ws
    _bg   = "#0066FF" if _active else ("#ECFDF5" if _done else "#F8FAFF")
    _col  = "#fff"    if _active else ("#10B981" if _done else "#94A3B8")
    _fw   = "800"     if _active else "600"
    _prefix = "✓ " if _done else ""
    _step_html += (
        f'<div style="flex:1;text-align:center;padding:14px 8px;background:{_bg};'
        f'color:{_col};font-weight:{_fw};font-size:0.875rem;border-right:1px solid #E2E8F0;">'
        f'<span style="font-size:1rem;font-weight:800;">{_prefix}Paso {_sn}</span><br>'
        f'<span style="font-size:0.8125rem;opacity:0.85;">{_sl}</span></div>'
    )
_step_html += '</div>'
st.markdown(_step_html, unsafe_allow_html=True)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PASO 1 — SUBIR FACTURAS + CONECTAR ALEGRA
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

if st.session_state.wizard_step == 1:

    col_upload, col_alegra = st.columns([3, 2], gap="large")

    # ── Sección de upload ─────────────────────────────────────────────────
    with col_upload:
        st.markdown("""
        <div class="step-header">
            <div class="step-num">1</div>
            <h2>📄 Sube tus facturas</h2>
        </div>
        """, unsafe_allow_html=True)
        st.caption("PDF, PNG, JPG, JPEG o WEBP · Máximo 50 archivos")

        uploaded_files = st.file_uploader(
            "Arrastra tus archivos aquí",
            type=["pdf", "jpg", "jpeg", "png", "webp"],
            accept_multiple_files=True,
            label_visibility="collapsed",
        )

        if uploaded_files:
            if len(uploaded_files) > 50:
                st.warning("Máximo 50 archivos. Se procesarán los primeros 50.")
                uploaded_files = uploaded_files[:50]

            st.markdown(f"**{len(uploaded_files)} archivo(s) seleccionado(s)**")
            _cols = st.columns(min(len(uploaded_files), 4))
            for _i, _uf in enumerate(uploaded_files):
                with _cols[_i % 4]:
                    _ext = Path(_uf.name).suffix.upper().replace(".", "")
                    st.markdown(
                        f'<div style="background:#F8FAFF;border:1px solid #E2E8F0;'
                        f'border-radius:8px;padding:6px 10px;margin:3px 0;font-size:0.8rem;">'
                        f'<strong style="color:#0066FF;">{_ext}</strong> {_uf.name}</div>',
                        unsafe_allow_html=True,
                    )

            st.markdown("")
            if st.button("⚙️ Procesar con IA", type="primary", use_container_width=True):
                _results, _names, _errors = [], [], []
                _bar = st.progress(0)
                _status = st.empty()
                _total  = len(uploaded_files)

                for _idx, _uf in enumerate(uploaded_files, 1):
                    _status.markdown(f"🧠 Procesando **{_idx}/{_total}**: `{_uf.name}`...")
                    try:
                        _suf = Path(_uf.name).suffix.lower()
                        with tempfile.NamedTemporaryFile(delete=False, suffix=_suf) as _tmp:
                            _tmp.write(_uf.getbuffer())
                            _tmp_path = _tmp.name
                        _datos = process_pdf(_tmp_path) if _suf == ".pdf" else process_image(_tmp_path)
                        _results.append(_datos)
                        _names.append(_uf.name)
                    except Exception as _e:
                        _errors.append((_uf.name, str(_e)))
                    _bar.progress(_idx / _total)

                _status.empty()
                _bar.empty()

                if _results:
                    # Limpiar caché de sugerencias IA de lote anterior
                    for _ck in [k for k in st.session_state if k.startswith(("ai_suggest_", "apply_all_account_", "apply_all_tax_"))]:
                        del st.session_state[_ck]
                    st.session_state.batch_results = _results
                    st.session_state.batch_names   = _names
                    st.session_state.wizard_step   = 2
                    st.session_state.reviewed_invoices = []
                    st.session_state.upload_results    = []
                    if _errors:
                        for _n, _e in _errors:
                            st.warning(f"No se pudo procesar {_n}: {_e}")
                    st.rerun()
                else:
                    for _n, _e in _errors:
                        st.error(f"Error en {_n}: {_e}")

    # ── Sección Alegra ────────────────────────────────────────────────────
    with col_alegra:
        st.markdown("""
        <div class="step-header">
            <div class="step-num" style="background:#10B981;">A</div>
            <h2>🔗 Alegra <span style="font-size:0.875rem;font-weight:500;color:#94A3B8;">(opcional)</span></h2>
        </div>
        """, unsafe_allow_html=True)

        if not _ALEGRA_OK:
            st.warning("Módulo alegra_client no disponible.")
        elif st.session_state.alegra_connected:
            st.success(f"**Conectado** como `{st.session_state.alegra_email}`")
            _cats = st.session_state.alegra_catalogs
            st.caption(
                f"{len(_cats['accounts'])} cuentas imputables · "
                f"{len(_cats['taxes'])} impuestos · "
                f"{len(_cats['cost_centers'])} centros de costo"
            )
            if st.button("Desconectar de Alegra", use_container_width=True, key="btn_disconnect_p1"):
                st.session_state.alegra_connected = False
                st.session_state.alegra_email     = ""
                st.session_state.alegra_token     = ""
                st.session_state.alegra_catalogs  = _defaults["alegra_catalogs"]
                st.rerun()
        else:
            st.caption("Conecta para habilitar subida directa de facturas de compra a Alegra.")
            _email = st.text_input(
                "Email Alegra", placeholder="usuario@empresa.co",
                value=st.session_state.alegra_email, key="inp_ale_email"
            )
            _token = st.text_input(
                "Token Alegra", placeholder="Token de API",
                type="password", value=st.session_state.alegra_token, key="inp_ale_token"
            )
            if st.button("Conectar a Alegra", type="primary", use_container_width=True, key="btn_connect_p1"):
                if not _email or not _token:
                    st.error("Ingresa email y token.")
                else:
                    with st.spinner("Verificando credenciales y descargando catálogos..."):
                        try:
                            _client = AlegraClient(_email, _token)
                            if not _client.ping():
                                st.error("Credenciales inválidas. Verifica tu email y token de Alegra.")
                            else:
                                _cats = _client.get_catalogs()
                                st.session_state.alegra_connected = True
                                st.session_state.alegra_email     = _email
                                st.session_state.alegra_token     = _token
                                st.session_state.alegra_catalogs  = _cats
                                st.rerun()
                        except AlegraAuthError:
                            st.error("Credenciales inválidas. Verifica tu email y token.")
                        except Exception as _e:
                            st.error(f"No se pudo conectar a Alegra: {_e}")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PASO 2 — REVISIÓN POR FACTURA
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

elif st.session_state.wizard_step == 2:

    st.markdown("""
    <div class="step-header">
        <div class="step-num">2</div>
        <h2>✏️ Revisa y parametriza cada factura</h2>
    </div>
    """, unsafe_allow_html=True)

    _has_alegra = _ALEGRA_OK and st.session_state.alegra_connected
    _accounts   = st.session_state.alegra_catalogs.get("accounts", [])
    _taxes      = st.session_state.alegra_catalogs.get("taxes", [])
    _ccs        = st.session_state.alegra_catalogs.get("cost_centers", [])

    if _has_alegra:
        _acc_opts = _account_options()
        _tax_opts = _tax_options()
        _cc_opts  = _cc_options()
        st.info(
            f"🔗 Alegra conectado · {len(_accounts)} cuentas · "
            f"{len(_taxes)} impuestos · {len(_ccs)} centros de costo"
        )
    else:
        st.caption("Alegra no conectado — solo revisión para Excel.")

    # Constantes para retenciones (globales por lote, por simplicidad UX)
    TIPOS = ["Compra de mercancía", "Gasto operacional", "Servicio", "Activo fijo"]
    PAGOS = ["Contado", "Crédito 30 días", "Crédito 60 días", "Crédito 90 días"]

    # Colectores de valores editados (llenados en el loop de expanders)
    _invoice_edits: list[dict] = []

    for _i, _datos in enumerate(st.session_state.batch_results):
        _prov_name = _datos.get("proveedor_nombre") or f"Factura {_i + 1}"
        _num_fac   = _datos.get("numero_factura") or "—"
        _nit       = str(_datos.get("proveedor_nit") or "")
        _archivo   = (
            st.session_state.batch_names[_i]
            if _i < len(st.session_state.batch_names) else ""
        )

        with st.expander(
            f"📄  {_i + 1}.  {_prov_name}  ·  {_num_fac}",
            expanded=(_i == 0),
        ):
            # ── Campos editables básicos ──────────────────────────────────
            _c1, _c2, _c3 = st.columns(3)
            with _c1:
                _e_prov  = st.text_input("Proveedor",       value=_prov_name,                  key=f"e_prov_{_i}")
                _e_nit   = st.text_input("NIT",             value=_nit,                         key=f"e_nit_{_i}")
            with _c2:
                _e_num   = st.text_input("# Factura",       value=_num_fac,                     key=f"e_num_{_i}")
                _e_fecha = st.text_input("Fecha emisión",   value=_datos.get("fecha_emision",""),  key=f"e_fecha_{_i}")
            with _c3:
                _e_vence = st.text_input("Fecha vence",     value=_datos.get("fecha_vencimiento",""), key=f"e_vence_{_i}")
                _e_total = st.number_input(
                    "Total a pagar",
                    value=float(_datos.get("total_a_pagar") or 0),
                    min_value=0.0, format="%.0f", key=f"e_tot_{_i}"
                )

            # ── Tabla de ítems con catálogos ─────────────────────────────
            _items_raw = _datos.get("items") or []
            if isinstance(_items_raw, str):
                try:
                    _items_raw = json.loads(_items_raw)
                except Exception:
                    _items_raw = []

            if not _items_raw:
                _items_raw = [{
                    "descripcion":    f"Factura {_num_fac} - {_prov_name}",
                    "cantidad":       1,
                    "valor_unitario": float(_datos.get("subtotal") or 0),
                    "valor_total":    float(_datos.get("subtotal") or 0),
                }]

            # Construir filas: memoria → sugerencia IA (cuenta + impuesto) → vacío
            _item_rows = []
            _apply_all_acc = st.session_state.get(f"apply_all_account_{_i}")
            _apply_all_tax = st.session_state.get(f"apply_all_tax_{_i}")

            for _item_idx, _item in enumerate(_items_raw):
                _desc = str(_item.get("descripcion") or "")
                _mem  = get_item_memory(_e_nit, _desc) if (_ALEGRA_OK and _e_nit) else {}

                # Obtener sugerencia IA (dict {account_id, tax_id}), cacheada por ítem
                _ai_result: dict = {}
                if _has_alegra and _acc_opts and _desc:
                    _ai_key = f"ai_suggest_{_i}_{_item_idx}"
                    if _ai_key not in st.session_state:
                        st.session_state[_ai_key] = suggest_account_for_item(
                            _desc, _e_prov, _accounts, _taxes
                        )
                    _ai_result = st.session_state[_ai_key] or {}

                # Cuenta: apply_all > memoria > IA
                _cuenta_label = _NONE_LABEL
                if _has_alegra and _acc_opts:
                    if _apply_all_acc and _apply_all_acc != _NONE_LABEL:
                        _cuenta_label = _apply_all_acc
                    else:
                        _cuenta_label = _account_id_to_label(_mem.get("cuenta_id"))
                        if _cuenta_label == _NONE_LABEL:
                            _cuenta_label = _account_id_to_label(_ai_result.get("account_id"))

                # Impuesto: apply_all > memoria > IA
                _tax_label = _NONE_LABEL
                if _has_alegra and len(_tax_opts) > 1:
                    if _apply_all_tax and _apply_all_tax != _NONE_LABEL:
                        _tax_label = _apply_all_tax
                    else:
                        _tax_label = _tax_id_to_label(_mem.get("impuesto_id"))
                        if _tax_label == _NONE_LABEL:
                            _tax_label = _tax_id_to_label(_ai_result.get("tax_id"))

                _item_rows.append({
                    "Descripción":      _desc,
                    "Precio":           float(_item.get("valor_total") or _item.get("valor_unitario") or 0),
                    "Cantidad":         float(_item.get("cantidad") or 1),
                    **({"Cuenta": _cuenta_label} if _has_alegra and _acc_opts else {}),
                    **({"Impuesto": _tax_label} if _has_alegra and len(_tax_opts) > 1 else {}),
                    **({"Centro de Costo": _cc_id_to_label(_mem.get("centro_costo_id"))} if _has_alegra and len(_cc_opts) > 1 else {}),
                })

            _df_items = pd.DataFrame(_item_rows)

            # Construir column_config dinámico
            _col_cfg: dict = {
                "Descripción": st.column_config.TextColumn("Descripción", width="large"),
                "Precio":      st.column_config.NumberColumn("Precio",    format="$ %,.0f", width="medium"),
                "Cantidad":    st.column_config.NumberColumn("Cant.",                        width="small"),
            }
            if _has_alegra and _acc_opts and "Cuenta" in _df_items.columns:
                _col_cfg["Cuenta"] = st.column_config.SelectboxColumn(
                    "Cuenta Alegra", options=_acc_opts, width="large",
                    help="Solo cuentas imputables (detalle). Las cuentas agrupadoras fueron filtradas."
                )
            if _has_alegra and len(_tax_opts) > 1 and "Impuesto" in _df_items.columns:
                _col_cfg["Impuesto"] = st.column_config.SelectboxColumn(
                    "Impuesto", options=_tax_opts, width="medium"
                )
            if _has_alegra and len(_cc_opts) > 1 and "Centro de Costo" in _df_items.columns:
                _col_cfg["Centro de Costo"] = st.column_config.SelectboxColumn(
                    "Centro de Costo", options=_cc_opts, width="medium"
                )

            st.caption("Tabla de ítems — editable. Asigna cuenta contable por ítem:")
            _edited_df = st.data_editor(
                _df_items,
                use_container_width=True,
                hide_index=True,
                num_rows="fixed",
                column_config=_col_cfg,
                key=f"items_editor_{_i}",
            )

            # ── Botón "Aplicar cuenta e impuesto del primer ítem a todos" ──
            if _has_alegra and _acc_opts and len(_item_rows) > 1 and "Cuenta" in _df_items.columns:
                if st.button(
                    "↓ Aplicar cuenta e impuesto del primer ítem a todos",
                    key=f"btn_apply_all_{_i}",
                    help="Copia la cuenta y el impuesto del primer ítem a todas las filas de esta factura",
                ):
                    _first_acc = str(_edited_df.iloc[0].get("Cuenta", _NONE_LABEL))
                    _first_tax = str(_edited_df.iloc[0].get("Impuesto", _NONE_LABEL)) if "Impuesto" in _edited_df.columns else _NONE_LABEL
                    st.session_state[f"apply_all_account_{_i}"] = _first_acc
                    st.session_state[f"apply_all_tax_{_i}"]     = _first_tax
                    st.rerun()

            # ── Retenciones por factura ───────────────────────────────────
            st.markdown("**Retenciones y parametrización:**")
            _rc1, _rc2, _rc3 = st.columns(3)
            with _rc1:
                _tipo_comp = st.selectbox("Tipo comprobante", TIPOS, key=f"tipo_{_i}")
            with _rc2:
                _forma_pago = st.selectbox("Forma de pago", PAGOS, key=f"fpago_{_i}")
            with _rc3:
                _centro = st.text_input("Centro de costo (PUC)", value="", placeholder="Ej: CC001", key=f"centro_{_i}")

            _sub = float(_datos.get("subtotal") or 0)
            _iva = float(_datos.get("valor_iva") or 0)
            _ret1, _ret2 = st.columns(2)
            with _ret1:
                _rf25  = st.checkbox("ReteFuente 2.5%",       key=f"rf25_{_i}")
                _rf35  = st.checkbox("ReteFuente 3.5% (no declarante)", key=f"rf35_{_i}")
            with _ret2:
                _riva  = st.checkbox("ReteIVA 15% del IVA",   key=f"riva_{_i}")
                _rcol, _rtcol = st.columns([3, 2])
                with _rcol:
                    _rica = st.checkbox("ReteICA",             key=f"rica_{_i}")
                with _rtcol:
                    _tasa_rica = st.number_input(
                        "Tasa ‰", min_value=0.0, max_value=20.0,
                        value=4.14, step=0.01, format="%.2f", key=f"trica_{_i}"
                    )

            # Guardar edits para este índice
            _invoice_edits.append({
                "idx":           _i,
                "proveedor":     _e_prov,
                "nit":           _e_nit,
                "numero":        _e_num,
                "fecha_emision": _e_fecha,
                "fecha_vence":   _e_vence,
                "total_a_pagar": _e_total,
                "subtotal":      _sub,
                "valor_iva":     _iva,
                "edited_df":     _edited_df,
                "tipo_comp":     _tipo_comp,
                "forma_pago":    _forma_pago,
                "centro":        _centro,
                "rf25":          _rf25,
                "rf35":          _rf35,
                "riva":          _riva,
                "rica":          _rica,
                "tasa_rica":     _tasa_rica,
                "_original":     _datos,
                "_archivo":      _archivo,
            })

    # ── Navegación ────────────────────────────────────────────────────────
    st.divider()
    _nav1, _nav2 = st.columns([1, 3])
    with _nav1:
        if st.button("← Volver al Paso 1", use_container_width=True):
            st.session_state.wizard_step = 1
            st.rerun()

    with _nav2:
        _lbl_btn2 = (
            f"Continuar a Sincronización → ({len(st.session_state.batch_results)} facturas)"
        )
        if st.button(_lbl_btn2, type="primary", use_container_width=True):
            # Procesar todos los edits y construir reviewed_invoices + historial
            _reviewed = []
            _historial_entries = []
            _ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            for _edit in _invoice_edits:
                _orig  = _edit["_original"]
                _nit_e = _edit["nit"]
                _sub_e = float(_edit["subtotal"])
                _iva_e = float(_edit["valor_iva"])

                # Convertir filas del data_editor a categorías Alegra
                _categories = []
                _items_detalle = []
                _df_ed = _edit["edited_df"]

                for _, _row in _df_ed.iterrows():
                    _desc     = str(_row.get("Descripción", ""))
                    _precio   = float(_row.get("Precio", 0))
                    _cant     = float(_row.get("Cantidad", 1))
                    _acc_lbl  = _row.get("Cuenta", _NONE_LABEL)
                    _tax_lbl  = _row.get("Impuesto", _NONE_LABEL)
                    _cc_lbl   = _row.get("Centro de Costo", _NONE_LABEL)
                    _acc_id   = _label_to_account_id(_acc_lbl) if _has_alegra else None
                    _tax_id   = _label_to_tax_id(_tax_lbl)     if _has_alegra else None
                    _cc_id    = _label_to_cc_id(_cc_lbl)        if _has_alegra else None

                    if _acc_id:
                        _categories.append({
                            "account_id":     _acc_id,
                            "price":          _precio,
                            "quantity":       int(_cant),
                            "observations":   f"{_desc} | Factura {_edit['numero']}",
                            "tax_id":         _tax_id,
                            "cost_center_id": _cc_id,
                        })

                    _items_detalle.append({
                        "Descripción": _desc,
                        "Tipo":        "Producto",
                        "Cuenta PUC":  str(_acc_id or ""),
                        "Cantidad":    _cant,
                        "Valor Total": _precio,
                        # Para memoria
                        "descripcion":     _desc,
                        "cuenta_id":       _acc_id,
                        "impuesto_id":     _tax_id,
                        "centro_costo_id": _cc_id,
                    })

                # Guardar memoria por NIT+ítem si hay cuenta asignada
                if _ALEGRA_OK and _nit_e and any(it.get("cuenta_id") for it in _items_detalle):
                    save_invoice_memory(_nit_e, _items_detalle)

                # Entrada reviewed
                _reviewed.append({
                    "proveedor_nombre":  _edit["proveedor"],
                    "proveedor_nit":     _nit_e,
                    "numero_factura":    _edit["numero"],
                    "fecha_emision":     _edit["fecha_emision"],
                    "fecha_vencimiento": _edit["fecha_vence"],
                    "subtotal":          _sub_e,
                    "total_a_pagar":     float(_edit["total_a_pagar"]),
                    "categories":        _categories,
                    "_archivo":          _edit["_archivo"],
                    "_original":         _orig,
                })

                # Entrada historial (para Excel)
                _historial_entries.append({
                    "proveedor":         _edit["proveedor"],
                    "nit_proveedor":     _nit_e,
                    "numero_factura":    _edit["numero"],
                    "fecha_emision":     _edit["fecha_emision"],
                    "fecha_vencimiento": _edit["fecha_vence"],
                    "subtotal":          _sub_e,
                    "pct_iva":           float(_orig.get("porcentaje_iva") or 0),
                    "valor_iva":         _iva_e,
                    "total_a_pagar":     float(_edit["total_a_pagar"]),
                    "total_bruto":       float(_orig.get("total_bruto") or _sub_e),
                    "direccion":         _orig.get("direccion_proveedor") or "",
                    "telefono":          _orig.get("telefono_proveedor")  or "",
                    "comprador":         _orig.get("comprador_nombre")    or "",
                    "nit_comprador":     _orig.get("comprador_nit")       or "",
                    "tipo_comprobante":  _edit["tipo_comp"],
                    "centro_costo":      _edit["centro"],
                    "forma_pago":        _edit["forma_pago"],
                    "aplica_rf25":       _edit["rf25"],
                    "rf25_valor":        round(_sub_e * 0.025) if _edit["rf25"] else 0,
                    "aplica_rf35":       _edit["rf35"],
                    "rf35_valor":        round(_sub_e * 0.035) if _edit["rf35"] else 0,
                    "aplica_riva":       _edit["riva"],
                    "riva_valor":        round(_iva_e * 0.15)  if _edit["riva"] else 0,
                    "aplica_rica":       _edit["rica"],
                    "tasa_rica":         _edit["tasa_rica"],
                    "rica_valor":        round(_sub_e * (_edit["tasa_rica"] / 1000)) if _edit["rica"] else 0,
                    "items_detalle":     [
                        {k: v for k, v in it.items()
                         if k in ("Descripción","Tipo","Cuenta PUC","Cantidad","Valor Total")}
                        for it in _items_detalle
                    ],
                    "_timestamp": _ts,
                    "_archivo":   _edit["_archivo"],
                })

            st.session_state.reviewed_invoices = _reviewed
            st.session_state.historial         = _historial_entries
            st.session_state.wizard_step       = 3
            st.rerun()

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PASO 3 — SINCRONIZACIÓN + EXCEL
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

elif st.session_state.wizard_step == 3:

    st.markdown("""
    <div class="step-header">
        <div class="step-num">3</div>
        <h2>🚀 Sincronizar y Exportar</h2>
    </div>
    """, unsafe_allow_html=True)

    _reviewed = st.session_state.reviewed_invoices
    _hist     = st.session_state.historial

    # ── Métricas ──────────────────────────────────────────────────────────
    _m1, _m2, _m3 = st.columns(3)
    with _m1:
        st.metric("📦 Facturas", len(_hist))
    with _m2:
        _tot = sum(h.get("total_a_pagar", 0) or 0 for h in _hist)
        st.metric("💰 Total a pagar", f"${_tot:,.0f}")
    with _m3:
        _iva = sum(h.get("valor_iva", 0) or 0 for h in _hist)
        st.metric("📋 Total IVA", f"${_iva:,.0f}")

    st.divider()

    # ── Vista previa ──────────────────────────────────────────────────────
    _PREVIEW_COLS = ["proveedor", "nit_proveedor", "numero_factura",
                     "fecha_emision", "subtotal", "total_a_pagar",
                     "tipo_comprobante", "forma_pago"]
    if _hist:
        _df_prev = pd.DataFrame(_hist)
        _df_prev = _df_prev[[c for c in _PREVIEW_COLS if c in _df_prev.columns]]
        st.markdown("**Vista previa del lote:**")
        st.dataframe(_df_prev, use_container_width=True, hide_index=True)

    st.divider()

    # ── Descarga Excel (siempre disponible) ───────────────────────────────
    st.markdown("### 📥 Exportar a Excel")
    if _hist:
        _exc_buf = crear_excel(_hist)
        _dc1, _dc2 = st.columns(2)
        with _dc1:
            st.markdown('<div class="download-btn">', unsafe_allow_html=True)
            st.download_button(
                label="📥 Descargar Excel (Resumen + Items)",
                data=_exc_buf,
                file_name=f"ContaFlow_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.markdown('</div>', unsafe_allow_html=True)
        with _dc2:
            _csv_data = _df_prev.to_csv(index=False, encoding="utf-8-sig")
            st.download_button(
                label="📥 Descargar CSV",
                data=_csv_data,
                file_name=f"ContaFlow_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv",
                use_container_width=True,
            )
    else:
        st.info("No hay facturas en el historial para exportar.")

    # ── Subida a Alegra ───────────────────────────────────────────────────
    if _ALEGRA_OK and st.session_state.alegra_connected and _reviewed:
        st.divider()
        st.markdown("### 🔗 Subir a Alegra")

        # Verificar que hay facturas con categorías asignadas
        _con_cats = [r for r in _reviewed if r.get("categories")]
        _sin_cats = [r for r in _reviewed if not r.get("categories")]

        if _sin_cats:
            st.warning(
                f"⚠️ {len(_sin_cats)} factura(s) sin cuenta contable asignada — "
                "no se subirán a Alegra. Vuelve al Paso 2 para asignar cuentas."
            )

        _results_already = st.session_state.upload_results

        if _results_already:
            # Mostrar resultados previos
            _ok_r   = [r for r in _results_already if r.get("ok")]
            _err_r  = [r for r in _results_already if not r.get("ok")]
            st.success(f"✅ {len(_ok_r)} subidas · ❌ {len(_err_r)} errores")

            for _r in _ok_r:
                st.markdown(
                    f'<div style="background:#ECFDF5;border-left:4px solid #10B981;'
                    f'padding:8px 14px;border-radius:8px;margin:4px 0;font-size:0.875rem;">'
                    f'✅ <strong>{_r["numero"]}</strong> — {_r["proveedor"]} '
                    f'→ Alegra ID <strong>{_r.get("alegra_id")}</strong></div>',
                    unsafe_allow_html=True,
                )
            for _r in _err_r:
                _retry_container = st.container()
                with _retry_container:
                    _ecol1, _ecol2 = st.columns([5, 1])
                    with _ecol1:
                        st.error(
                            f"❌ **{_r['numero']}** — {_r['proveedor']}: "
                            f"{_r.get('error', 'Error desconocido')}"
                        )
                    with _ecol2:
                        if st.button("Reintentar", key=f"retry_{_r['numero']}"):
                            # Reintentar solo esta factura
                            _target = next(
                                (rv for rv in _con_cats if rv["numero_factura"] == _r["numero"]),
                                None,
                            )
                            if _target:
                                with st.spinner(f"Reintentando {_r['numero']}..."):
                                    try:
                                        _client = AlegraClient(
                                            st.session_state.alegra_email,
                                            st.session_state.alegra_token,
                                        )
                                        _resp = _client.create_purchase_invoice(_target)
                                        # Actualizar resultado
                                        st.session_state.upload_results = [
                                            {**r, "ok": True, "alegra_id": _resp.get("id")}
                                            if r["numero"] == _r["numero"] else r
                                            for r in st.session_state.upload_results
                                        ]
                                        st.rerun()
                                    except (AlegraAuthError, AlegraAPIError) as _exc:
                                        st.error(friendly_error(_exc) if isinstance(_exc, AlegraAPIError) else str(_exc))
                                    except Exception as _exc:
                                        st.error(f"Error inesperado: {_exc}")

            if st.button("🔄 Resubir todo", use_container_width=True):
                st.session_state.upload_results = []
                st.rerun()

        elif _con_cats:
            st.caption(
                f"{len(_con_cats)} factura(s) listas para subir "
                f"({len(_sin_cats)} sin cuenta — se omitirán)."
            )
            if st.button(
                f"🚀 Subir {len(_con_cats)} factura(s) aprobadas a Alegra",
                type="primary",
                use_container_width=True,
            ):
                _upload_results = []
                _client = AlegraClient(
                    st.session_state.alegra_email,
                    st.session_state.alegra_token,
                )
                _prog = st.progress(0)
                _stat = st.empty()
                _total_up = len(_con_cats)

                for _ui, _inv in enumerate(_con_cats, 1):
                    _stat.markdown(
                        f"⬆️ Subiendo **{_ui}/{_total_up}**: "
                        f"`{_inv['numero_factura']}` — {_inv['proveedor_nombre']}..."
                    )
                    try:
                        _resp = _client.create_purchase_invoice(_inv)
                        _upload_results.append({
                            "ok":       True,
                            "numero":   _inv["numero_factura"],
                            "proveedor": _inv["proveedor_nombre"],
                            "alegra_id": _resp.get("id"),
                        })
                    except AlegraAuthError as _exc:
                        _upload_results.append({
                            "ok":       False,
                            "numero":   _inv["numero_factura"],
                            "proveedor": _inv["proveedor_nombre"],
                            "error":    "Sesión de Alegra expirada. Vuelve al Paso 1 y reconecta.",
                        })
                    except AlegraAPIError as _exc:
                        _upload_results.append({
                            "ok":       False,
                            "numero":   _inv["numero_factura"],
                            "proveedor": _inv["proveedor_nombre"],
                            "error":    friendly_error(_exc),
                        })
                    except Exception as _exc:
                        _upload_results.append({
                            "ok":       False,
                            "numero":   _inv["numero_factura"],
                            "proveedor": _inv["proveedor_nombre"],
                            "error":    f"Error inesperado. Intenta de nuevo.",
                        })
                    _prog.progress(_ui / _total_up)

                _stat.empty()
                _prog.empty()
                st.session_state.upload_results = _upload_results
                st.rerun()

    elif not st.session_state.alegra_connected:
        st.info("Conecta Alegra en el Paso 1 para habilitar la subida directa.")

    # ── Navegación ────────────────────────────────────────────────────────
    st.divider()
    _b1, _b2 = st.columns([1, 1])
    with _b1:
        if st.button("← Revisar de nuevo (Paso 2)", use_container_width=True):
            st.session_state.wizard_step   = 2
            st.session_state.upload_results = []
            st.rerun()
    with _b2:
        if st.button("🔄 Procesar nuevo lote (Paso 1)", use_container_width=True):
            for _k in ("wizard_step", "batch_results", "batch_names",
                       "reviewed_invoices", "upload_results"):
                st.session_state[_k] = _defaults[_k]
            st.rerun()
