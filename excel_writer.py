"""
excel_writer.py — Crea y actualiza el Excel de resultados con openpyxl

El Excel tiene una sola hoja ("Facturas") con los mismos campos que el CSV,
cabeceras en negrita con fondo azul oscuro y la primera fila fija (freeze).
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from extractor import CAMPOS_CSV

EXCEL_FILENAME = "ContaFlow_Facturas.xlsx"
SHEET_NAME = "Facturas"

# Etiquetas legibles para las cabeceras del Excel
HEADERS = {
    "numero_factura":    "# Factura",
    "proveedor_nombre":  "Proveedor",
    "proveedor_nit":     "NIT Proveedor",
    "comprador_nombre":  "Comprador",
    "comprador_nit":     "NIT Comprador",
    "fecha_emision":     "Fecha Emision",
    "fecha_vencimiento": "Fecha Vencimiento",
    "forma_pago":        "Forma de Pago",
    "subtotal":          "Subtotal",
    "porcentaje_iva":    "% IVA",
    "valor_iva":         "Valor IVA",
    "total_factura":     "Total Factura",
    "es_autorretenedor": "Autorretenedor",
    "aplica_retefuente": "Retefuente",
    "_archivo_fuente":   "Archivo PDF",
    "_procesado_en":     "Procesado En",
}

_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HEADER_FILL  = PatternFill(fill_type="solid", fgColor="1F4E79")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _build_header_row(ws):
    for col_idx, field in enumerate(CAMPOS_CSV, start=1):
        cell = ws.cell(row=1, column=col_idx, value=HEADERS.get(field, field))
        cell.font  = _HEADER_FONT
        cell.fill  = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def _autofit_columns(ws):
    """Ajusta el ancho de cada columna al contenido mas largo."""
    for col_idx in range(1, len(CAMPOS_CSV) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)


def create_workbook() -> openpyxl.Workbook:
    """Crea un libro nuevo con la hoja 'Facturas' y sus cabeceras."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _build_header_row(ws)
    return wb


def load_or_create_workbook(local_path: str) -> openpyxl.Workbook:
    """Carga el Excel si existe; si no, crea uno nuevo."""
    path = Path(local_path)
    if path.exists():
        wb = openpyxl.load_workbook(local_path)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME, 0)
            _build_header_row(ws)
        return wb
    return create_workbook()


def append_records(wb: openpyxl.Workbook, records: list[dict]):
    """Agrega filas al final de la hoja y reajusta los anchos de columna."""
    ws = wb[SHEET_NAME]
    for record in records:
        ws.append([record.get(field) for field in CAMPOS_CSV])
    _autofit_columns(ws)


def save_workbook(wb: openpyxl.Workbook, local_path: str):
    Path(local_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(local_path)
